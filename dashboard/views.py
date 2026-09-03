import csv, io, json, secrets, string, math
from datetime import date, timedelta
from django.contrib import messages
from django.conf import settings
from django.contrib.auth import authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Avg, Count, Max, Q
from django.http import HttpResponse, JsonResponse
from django.db.models.functions import TruncDate
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.text import slugify
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_exempt

from accounts.forms import PersonelOlusturForm, RolAtamaKuraliForm, RolForm
from accounts.models import PersonelProfili, Rol, RolAtamaKurali
from accounts.saha_blueprint import SAHA_BLUEPRINTS
from accounts.saha_services import create_missing_field_teams, create_missing_coordinators, create_missing_organization
from adres.forms import AdresCsvForm
from adres.models import Ilce, Mahalle, Yol
from adres.geocoding import talep_konumla
from talepler.forms import IsAltTuruForm, IsTuruForm, TalepForm, AboneForm
from talepler.models import GeriBildirim, IsTuru, IsAltTuru, IslemLogu, Talep, VatandasAramaKaydi, Abone, IsEmri, IsEmriFotograf, MobilToken, MobilBildirim
from talepler.services import akisa_yaz, koordinator_bul, kullanici_talepleri, log_yaz, talep_erisim_var_mi, uygun_saha_personelleri, bolge_saha_personelleri
from .forms import TalepFiltreForm
from .permissions import get_profile, panel_required, operational_profile


def _web_only_abone_role(rol):
    """Sayaç/abone operasyon rolü masaüstü web kanalında çalışır."""
    if not rol:
        return False
    kod=(rol.kod or "").casefold().replace("_","-")
    ad=(rol.ad or "").casefold()
    return kod in {"abone-sayac-saha","sayac-abone-saha"} or ("sayaç" in ad and "abone" in ad) or ("sayac" in ad and "abone" in ad)


@login_required
def home(request):
    if request.user.is_superuser:
        return redirect("dashboard:sistem")
    p=get_profile(request.user)
    if not p:
        raise PermissionDenied("Bu kullanıcıya personel profili tanımlanmamış.")

    # V56: abonelik personeli saha rolünden ayrıldı; ayrı Web/PC operasyon rolüdür.
    if p.rol.panel_tipi=="abone":
        return redirect("abonelik:izleme")

    # V29: yönlendirme kullanıcı adına göre değil rol + çalışma kanalı verisine göre yapılır.
    if p.rol.panel_tipi=="saha":
        # Sayaç/Abone rolü saha panel tipini kullansa da masaüstü/web işidir.
        # Veritabanında eski bir "mobil" değeri kalmış olsa bile mobil ekrana düşmez.
        if _web_only_abone_role(p.rol):
            return redirect("dashboard:abone_islemleri")
        if p.rol.calisma_kanali=="mobil":
            return redirect("dashboard:mobil_saha")
        return redirect("dashboard:saha")

    hedef={
        "admin":"dashboard:sistem",
        "185":"dashboard:cagri_185",
        "sef":"dashboard:sef",
        "merkez_ambar":"abonelik:ambar_yonetimi",
    }.get(p.rol.panel_tipi)
    if not hedef:
        raise PermissionDenied("Bu rol için giriş ekranı tanımlanmamış.")
    return redirect(hedef)



def _isemri_durum_from_talep(durum):
    return {
        "sahaya_atandi":"atandi",
        "kabul_edildi":"kabul_edildi",
        "yolda":"yolda",
        "yerinde":"yerinde",
        "islemde":"islemde",
        "onay_bekliyor":"onay_bekliyor",
        "tamamlandi":"tamamlandi",
        "iptal":"iptal",
    }.get(durum,"atandi")


def _is_emri_esitle(talep, kullanici=None, gonderen_birim=None):
    """Talep operasyonuyla tek iş emrini aynı anda güncel tut."""
    if not talep.sorumlu_saha_id:
        return None
    defaults={
        "gonderen_birim":gonderen_birim or (
            talep.sorumlu_koordinator.rol.ad
            if talep.sorumlu_koordinator_id and talep.sorumlu_koordinator.rol_id
            else f"{talep.is_turu.ad} Birimi"
        ),
        "olusturan":kullanici or talep.olusturan,
        "atanan_saha":talep.sorumlu_saha,
        "durum":_isemri_durum_from_talep(talep.durum),
        "atama_tarihi":timezone.now(),
    }
    emri,created=IsEmri.objects.get_or_create(talep=talep,defaults=defaults)
    emri.atanan_saha=talep.sorumlu_saha
    if gonderen_birim:
        emri.gonderen_birim=gonderen_birim
    elif not emri.gonderen_birim:
        emri.gonderen_birim=defaults["gonderen_birim"]
    emri.durum=_isemri_durum_from_talep(talep.durum)
    now=timezone.now()
    if talep.durum=="kabul_edildi" and not emri.kabul_tarihi: emri.kabul_tarihi=now
    if talep.durum=="yolda" and not emri.yola_cikis_tarihi: emri.yola_cikis_tarihi=now
    if talep.durum=="yerinde" and not emri.adrese_ulasma_tarihi: emri.adrese_ulasma_tarihi=now
    if talep.durum=="islemde" and not emri.mudahale_baslama_tarihi: emri.mudahale_baslama_tarihi=now
    if talep.durum=="onay_bekliyor":
        emri.saha_tamam_tarihi=talep.saha_tamam_bildirim_tarihi or now
        emri.sonuc_notu=talep.saha_sonuc_notu or emri.sonuc_notu
    if talep.durum=="tamamlandi":
        emri.sef_onay_tarihi=talep.sef_onay_tarihi or now
        emri.sonuc_notu=talep.saha_sonuc_notu or emri.sonuc_notu
    emri.save()
    return emri


GPS_SAHA_ESIGI_M=500


def _mesafe_m(lat1,lng1,lat2,lng2):
    """Haversine; iki koordinat arasındaki yaklaşık metre."""
    try:
        lat1,lng1,lat2,lng2=map(float,(lat1,lng1,lat2,lng2))
    except (TypeError,ValueError):
        return None
    r=6371000.0
    p1,p2=math.radians(lat1),math.radians(lat2)
    dp=math.radians(lat2-lat1)
    dl=math.radians(lng2-lng1)
    a=math.sin(dp/2)**2+math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return int(round(2*r*math.atan2(math.sqrt(a),math.sqrt(1-a))))


def _mobil_saha_gps_dogrula(request,talep,emri,aksiyon_adi):
    """V46 — Konum gerektiren saha adımlarında anlık GPS'i doğrular ve kaydeder."""
    raw_lat=(request.POST.get("gps_lat") or "").strip()
    raw_lng=(request.POST.get("gps_lng") or "").strip()
    if not raw_lat or not raw_lng:
        return False,None,f"{aksiyon_adi} için cihazın anlık konumu gerekli."
    try:
        gps_lat=float(raw_lat); gps_lng=float(raw_lng)
    except ValueError:
        return False,None,"Cihaz konumu okunamadı. Konum iznini kontrol edin."

    if talep.lat is None or talep.lng is None:
        return False,None,"Bu iş emrinin kayıtlı arıza koordinatı yok. Konum doğrulanmadan işlem ilerletilemez."

    mesafe=_mesafe_m(gps_lat,gps_lng,talep.lat,talep.lng)
    dogrulandi=mesafe is not None and mesafe<=GPS_SAHA_ESIGI_M

    emri.gps_lat=raw_lat
    emri.gps_lng=raw_lng
    emri.gps_mesafe_m=mesafe
    emri.gps_dogrulandi=dogrulandi
    emri.gps_dogrulama_tarihi=timezone.now()
    emri.save(update_fields=[
        "gps_lat","gps_lng","gps_mesafe_m","gps_dogrulandi","gps_dogrulama_tarihi","guncellenme_tarihi"
    ])

    if not dogrulandi:
        mesafe_text=f"yaklaşık {mesafe} metre" if mesafe is not None else "hesaplanamayan bir mesafe"
        return False,mesafe,(
            f"GPS doğrulaması başarısız: iş emri konumuna {mesafe_text} uzaktasınız. "
            f"{GPS_SAHA_ESIGI_M} metre içine gelmeden '{aksiyon_adi}' işlemi tamamlanamaz."
        )
    return True,mesafe,None


def _telefon_rakam(telefon):
    return "".join(ch for ch in (telefon or "") if ch.isdigit())


def _talep_surec_adimlari(talep):
    """Talep detayında yaşam döngüsünü tek bakışta göstermek için görünüm verisi."""
    saha_operasyonu_bitti = talep.durum in ("onay_bekliyor", "tamamlandi")
    sef_onayi_bitti = bool(talep.sef_onay_tarihi) or talep.durum == "tamamlandi"
    vatandas_bitti = talep.vatandas_bildirim_durumu == "bilgilendirildi"

    adimlar = [
        {"kod":"185", "ad":"185 Kaydı", "tamam":True, "detay":timezone.localtime(talep.olusturulma_tarihi).strftime("%d.%m.%Y %H:%M")},
        {"kod":"koordinator", "ad":"Koordinatör", "tamam":bool(talep.sorumlu_koordinator_id) or talep.durum != "yeni", "detay":talep.sorumlu_koordinator.kullanici.username if talep.sorumlu_koordinator_id else "Yönlendirme bekleniyor"},
        {"kod":"saha", "ad":"Saha Atama", "tamam":bool(talep.sorumlu_saha_id), "detay":talep.sorumlu_saha.kullanici.username if talep.sorumlu_saha_id else "Atama bekleniyor"},
        {"kod":"operasyon", "ad":"Saha Operasyonu", "tamam":saha_operasyonu_bitti, "detay":"Saha tamamladı" if saha_operasyonu_bitti else (talep.get_durum_display() if talep.sorumlu_saha_id else "Saha bekleniyor")},
        {"kod":"onay", "ad":"Şef Onayı", "tamam":sef_onayi_bitti, "detay":timezone.localtime(talep.sef_onay_tarihi).strftime("%d.%m.%Y %H:%M") if talep.sef_onay_tarihi else "Onay bekleniyor"},
        {"kod":"vatandas", "ad":"Vatandaş Dönüşü", "tamam":vatandas_bitti, "detay":timezone.localtime(talep.vatandas_bildirim_tarihi).strftime("%d.%m.%Y %H:%M") if talep.vatandas_bildirim_tarihi else talep.get_vatandas_bildirim_durumu_display()},
    ]
    aktif_bulundu = False
    for adim in adimlar:
        adim["aktif"] = False
        if not adim["tamam"] and not aktif_bulundu:
            adim["aktif"] = True
            aktif_bulundu = True
    return adimlar


def _saha_foto_dogrula(upload):
    if not upload:
        return False,"Fotoğraf seçilmedi."
    content_type=(getattr(upload,"content_type","") or "").lower()
    if not content_type.startswith("image/"):
        return False,"Yalnızca fotoğraf dosyası yükleyebilirsiniz."
    if getattr(upload,"size",0)>8*1024*1024:
        return False,"Fotoğraf en fazla 8 MB olabilir."
    return True,""


def _is_emri_foto_slotlari(emri):
    """İş alt türünün V42 fotoğraf kuralını iş emrindeki yüklemelerle birleştirir."""
    if not emri or not emri.talep_id or not emri.talep.is_alt_turu_id:
        return []
    etiketler=emri.talep.is_alt_turu.zorunlu_foto_etiketleri
    mevcut={x.sira:x for x in emri.fotograflar.all()}
    return [
        {"sira":i,"etiket":etiket,"fotograf":mevcut.get(i),"tamam":i in mevcut}
        for i,etiket in enumerate(etiketler,start=1)
    ]


def _is_emri_foto_durumu(emri):
    slotlar=_is_emri_foto_slotlari(emri)
    yuklenen=sum(1 for x in slotlar if x["tamam"])
    return slotlar,yuklenen,len(slotlar)


def _mobil_bildirim_olustur(kullanici,is_emri,tip,baslik,mesaj):
    if not kullanici:
        return None
    return MobilBildirim.objects.create(
        kullanici=kullanici,
        is_emri=is_emri,
        tip=tip,
        baslik=baslik[:140],
        mesaj=mesaj[:400],
    )



def _saha_atamasini_yap(request,talep,saha,atan_kisi):
    """Şef/koordinatörün manuel seçtiği saha hesabına iş emri oluşturur."""
    talep.sorumlu_saha=saha
    talep.durum="sahaya_atandi"
    talep.save(update_fields=["sorumlu_saha","durum","guncellenme_tarihi"])

    is_emri=_is_emri_esitle(
        talep,
        request.user,
        atan_kisi.rol.ad if atan_kisi and atan_kisi.rol_id else f"{talep.is_turu.ad} Birimi",
    )
    ad=saha.kullanici.get_full_name() or saha.kullanici.username
    mesaj=(
        f"Talep {talep.ilce.ad} bölgesindeki {saha.rol.ad} ekibine ({ad}) "
        f"şef/koordinatör tarafından manuel atandı. İş emri: {is_emri.is_emri_no}."
    )
    akisa_yaz(talep,request.user,mesaj,talep.durum,request,True,"ASSIGN")
    log_yaz(
        request,"ASSIGN",mesaj,talep,"IsEmri",is_emri.pk,"","sahaya_atandi",
    )

    tip="acil" if talep.oncelik=="acil" else "yeni_is"
    baslik=("ACİL • " if talep.oncelik=="acil" else "")+f"Yeni İş Emri {is_emri.is_emri_no}"
    _mobil_bildirim_olustur(
        saha.kullanici,is_emri,tip,baslik,
        f"{talep.ilce.ad} / {talep.mahalle.ad} • {talep.is_turu.ad} → {talep.is_alt_turu.ad}",
    )
    return is_emri


# ---------------------------------------------------------------------
# V27 — Öncelik Bazlı İç Operasyon SLA
#
# İSU'nun yayımlanan kamu hizmet standardında içmesuyu / kanalizasyon
# arızaları için toplam tamamlanma süresi 2–8 saat aralığındadır.
# Aşağıdaki değerler resmi kurum SLA'sı iddiası değildir; bu proje içinde
# koordinatörün talebi sahaya yönlendirmesini hızlandırmak için kullanılan
# iç operasyon hedefleridir.
# ---------------------------------------------------------------------
SLA_ATAMA_DAKIKA = {
    "acil": 30,
    "yuksek": 60,
    "normal": 120,
    "dusuk": 240,
}


def _sla_atama_baslangici(talep):
    """Talebin şef/koordinatör kuyruğuna ilk düştüğü zamanı bul."""
    hareket = GeriBildirim.objects.filter(
        talep=talep,
        durum="sefe_gonderildi",
    ).order_by("tarih").first()
    return hareket.tarih if hareket else talep.olusturulma_tarihi


def _dakika_kisa_label(dakika):
    dakika=max(0,int(dakika or 0))
    if dakika >= 1440:
        gun, kalan=divmod(dakika,1440)
        saat=kalan//60
        return f"{gun}g {saat}s" if saat else f"{gun}g"
    if dakika >= 60:
        saat, dk=divmod(dakika,60)
        return f"{saat}s {dk}dk" if dk else f"{saat}s"
    return f"{dakika} dk"


def _sla_bilgisi(talep, now=None):
    """
    Sadece henüz sahaya atanmamış talepler için atama SLA bilgisini üretir.
    durum:
      uygun      -> sürenin %75'inden az kullanılmış
      yaklasiyor -> sürenin %75+ kısmı kullanılmış
      asildi     -> hedef süre geçmiş
      tamam      -> sahaya atanmış / SLA takibi bitmiş
    """
    hedef_dakika=SLA_ATAMA_DAKIKA.get(talep.oncelik,120)

    if talep.sorumlu_saha_id or talep.durum not in ("yeni","sefe_gonderildi"):
        return {
            "durum":"tamam",
            "hedef_dakika":hedef_dakika,
            "hedef_label":f"{hedef_dakika} dk" if hedef_dakika < 60 else f"{hedef_dakika//60} sa",
            "kalan_dakika":0,
            "asma_dakika":0,
            "kalan_label":"0 dk",
            "asma_label":"0 dk",
            "gecen_dakika":0,
            "yuzde":100,
            "baslangic":None,
            "son_tarih":None,
        }

    now=now or timezone.now()
    baslangic=_sla_atama_baslangici(talep)
    son_tarih=baslangic+timedelta(minutes=hedef_dakika)
    gecen=max(0,int((now-baslangic).total_seconds()//60))
    kalan=hedef_dakika-gecen
    yuzde=max(0,min(100,int((gecen/hedef_dakika)*100))) if hedef_dakika else 100

    if kalan < 0:
        durum="asildi"
    elif yuzde >= 75:
        durum="yaklasiyor"
    else:
        durum="uygun"

    if hedef_dakika < 60:
        hedef_label=f"{hedef_dakika} dk"
    elif hedef_dakika % 60 == 0:
        hedef_label=f"{hedef_dakika//60} sa"
    else:
        hedef_label=f"{hedef_dakika} dk"

    return {
        "durum":durum,
        "hedef_dakika":hedef_dakika,
        "hedef_label":hedef_label,
        "kalan_dakika":kalan,
        "asma_dakika":max(0,-kalan),
        "kalan_label":_dakika_kisa_label(max(0,kalan)),
        "asma_label":_dakika_kisa_label(max(0,-kalan)),
        "gecen_dakika":gecen,
        "yuzde":yuzde,
        "baslangic":baslangic,
        "son_tarih":son_tarih,
    }


def _sla_ekle(talepler):
    """Template kullanımına uygun SLA alanlarını Talep nesnelerine ekle."""
    now=timezone.now()
    sonuc=[]
    for talep in talepler:
        info=_sla_bilgisi(talep,now)
        talep.sla=info
        sonuc.append(talep)
    return sonuc


@panel_required("185","admin")
def cagri_185(request):
    operational_profile(request,"185")
    if request.method=="POST":
        form=TalepForm(request.POST)
        if form.is_valid():
            talep=form.save(commit=False)
            talep.olusturan=request.user
            abone_no=(form.cleaned_data.get("abone_no") or "").strip()
            if abone_no:
                talep.abone=Abone.objects.filter(abone_no__iexact=abone_no,aktif=True).first()

            # Listede yol yoksa 185 personeli gerçek cadde/sokak adını yazar.
            # İki alan da boşsa TalepForm.clean() zaten kaydı engeller.
            if not talep.yol_id:
                yol_adi=(form.cleaned_data.get("yol_serbest") or "").strip()
                yol_lower=yol_adi.casefold()
                if "bulvar" in yol_lower:
                    yol_turu="bulvar"
                elif "cadde" in yol_lower or "caddesi" in yol_lower:
                    yol_turu="cadde"
                elif "sokak" in yol_lower or "sokağı" in yol_lower or "sokagi" in yol_lower:
                    yol_turu="sokak"
                else:
                    yol_turu="diger"
                talep.yol,_=Yol.objects.get_or_create(
                    mahalle=talep.mahalle,
                    ad=yol_adi,
                    defaults={"tur":yol_turu,"aktif":True},
                )

            # Harita konumu: kullanıcı haritadan işaretlemediyse gerçek adresten
            # otomatik bulunur; servis ulaşılamazsa ilçe/yerel merkez kullanılır.
            talep.save()
            talep_konumla(talep, force=False, save=True)

            # 1) Talep oluşturma hareketi sistem tarafından otomatik yazılır.
            akisa_yaz(
                talep,request.user,
                "Talep 185 Çağrı Merkezi tarafından oluşturuldu.",
                "yeni",request,True,"CREATE"
            )

            # 2) Koordinatör yönlendirmesi de ayrı bir otomatik hareket olarak yazılır.
            talep.sorumlu_koordinator=koordinator_bul(talep)
            if talep.sorumlu_koordinator:
                talep.durum="sefe_gonderildi"
                talep.save(update_fields=["sorumlu_koordinator","durum","guncellenme_tarihi"])
                ad=talep.sorumlu_koordinator.kullanici.get_full_name() or talep.sorumlu_koordinator.kullanici.username
                akisa_yaz(
                    talep,request.user,
                    f"Talep {talep.ilce.ad} için {ad} koordinatörüne otomatik yönlendirildi.",
                    talep.durum,request,True,"ROUTE"
                )
            else:
                akisa_yaz(
                    talep,request.user,
                    "Bu ilçe/iş türü için uygun koordinatör bulunamadı; yönetici kontrolü bekleniyor.",
                    talep.durum,request,True,"ROUTE_WARNING"
                )

            # PRG (Post/Redirect/Get): kayıt başarıyla bittiğinde POST formu taşınmaz.
            # Böylece tüm alanlar otomatik olarak temizlenir ve yeni çağrı formu açılır.
            messages.success(
                request,
                f"{talep.talep_no} başarıyla oluşturuldu ve yönlendirildi. Yeni çağrı için form temizlendi."
            )
            return redirect(f"{reverse('dashboard:cagri_185')}?created={talep.pk}")
    else:
        form=TalepForm()

    qs=Talep.objects.select_related(
        "ilce","mahalle","yol","is_turu","is_alt_turu",
        "sorumlu_saha__kullanici","sorumlu_saha__rol",
        "sef_onaylayan","vatandas_bildirim_yapan"
    ).prefetch_related("vatandas_arama_kayitlari","geri_bildirimler").order_by("-olusturulma_tarihi")

    geri_bildirim_qs=qs.filter(
        durum="tamamlandi",
        vatandas_bildirim_durumu__in=["bekliyor","tekrar_aranacak"],
    ).order_by(
        "vatandas_bildirim_durumu",
        "-sef_onay_tarihi",
    )

    # 185 ekranındaki sol arama kuyruğu + sağ detay alanı için yaşam döngüsü
    # zamanlarını hazırla. Veriyi değiştirmez; mevcut otomatik hareket kayıtlarını okur.
    geri_bildirim_talepleri=list(geri_bildirim_qs[:50])
    for _talep in geri_bildirim_talepleri:
        _hareketler=sorted(list(_talep.geri_bildirimler.all()),key=lambda x:x.tarih)

        def _ilk_hareket(durum):
            return next((x for x in _hareketler if x.durum==durum),None)

        _talep.timeline_koordinator=_ilk_hareket("sefe_gonderildi")
        _talep.timeline_saha_atama=_ilk_hareket("sahaya_atandi")
        _talep.timeline_kabul=_ilk_hareket("kabul_edildi")
        _talep.timeline_yolda=_ilk_hareket("yolda")
        _talep.timeline_yerinde=_ilk_hareket("yerinde")
        _talep.timeline_mudahale=_ilk_hareket("islemde")
        _talep.surec_adimlari=_talep_surec_adimlari(_talep)
        _talep.son_arama_kaydi=next(iter(_talep.vatandas_arama_kayitlari.all()),None)

        # V15: bir vatandaş görüşmesi tek tıkla "yapıldı" sayılamaz.
        # Aramayı Başlat olayı IslemLogu'na kullanıcı/IP/saat ile yazılır.
        _son_baslatma=IslemLogu.objects.filter(
            talep=_talep,islem="CITIZEN_CALL_START"
        ).select_related("kullanici").order_by("-tarih").first()
        _son_sonuc=IslemLogu.objects.filter(
            talep=_talep,islem="CITIZEN_CALLBACK"
        ).order_by("-tarih").first()

        _talep.acik_arama_baslangici=None
        _talep.acik_arama_benim=False
        _talep.acik_arama_kullanici=None
        if _son_baslatma and (not _son_sonuc or _son_baslatma.tarih > _son_sonuc.tarih):
            _talep.acik_arama_baslangici=_son_baslatma.tarih
            _talep.acik_arama_kullanici=_son_baslatma.kullanici
            _talep.acik_arama_benim=(
                _son_baslatma.kullanici_id == request.user.id
            )

    son_bilgilendirilenler=qs.filter(
        durum="tamamlandi",
        vatandas_bildirim_durumu="bilgilendirildi",
    ).order_by("-vatandas_bildirim_tarihi")[:50]

    stats={
        "toplam":qs.count(),
        "yeni":qs.filter(durum__in=["yeni","sefe_gonderildi"]).count(),
        "islemde":qs.filter(durum__in=["sahaya_atandi","kabul_edildi","yolda","yerinde","islemde"]).count(),
        "onay":qs.filter(durum="onay_bekliyor").count(),
        "tamam":qs.filter(durum="tamamlandi").count(),
        "geri_bekleyen":geri_bildirim_qs.filter(vatandas_bildirim_durumu="bekliyor").count(),
        "geri_tekrar":geri_bildirim_qs.filter(vatandas_bildirim_durumu="tekrar_aranacak").count(),
        "geri_toplam":geri_bildirim_qs.count(),
        "geri_bugun":qs.filter(
            vatandas_bildirim_durumu="bilgilendirildi",
            vatandas_bildirim_tarihi__date=timezone.localdate(),
        ).count(),
    }
    return render(request,"dashboard/cagri_185.html",{
        "form":form,
        "son_talepler":qs[:50],
        "geri_bildirim_talepleri":geri_bildirim_talepleri,
        "son_bilgilendirilenler":son_bilgilendirilenler,
        "stats":stats,
    })

@login_required
@require_GET
def benzer_acik_talepler(request):
    """185 formunda benzer açık kayıtları uyarı amaçlı getirir; kaydı engellemez."""
    operational_profile(request,"185")
    telefon=_telefon_rakam(request.GET.get("telefon"))
    ilce_id=request.GET.get("ilce")
    mahalle_id=request.GET.get("mahalle")
    is_turu_id=request.GET.get("is_turu")
    is_alt_turu_id=request.GET.get("is_alt_turu")
    yol_id=request.GET.get("yol")
    yol_serbest=(request.GET.get("yol_serbest") or "").strip().casefold()

    if not (telefon and ilce_id and mahalle_id and is_turu_id):
        return JsonResponse({"ok":True,"count":0,"results":[]})

    qs=Talep.objects.select_related("ilce","mahalle","yol","is_turu","is_alt_turu").exclude(
        durum__in=["tamamlandi","iptal"]
    ).filter(ilce_id=ilce_id,mahalle_id=mahalle_id,is_turu_id=is_turu_id).order_by("-olusturulma_tarihi")

    if is_alt_turu_id:
        qs=qs.filter(is_alt_turu_id=is_alt_turu_id)
    if yol_id:
        qs=qs.filter(yol_id=yol_id)

    results=[]
    for t in qs[:60]:
        if _telefon_rakam(t.telefon) != telefon:
            continue
        if yol_serbest and not yol_id and t.yol_id and t.yol.ad.strip().casefold() != yol_serbest:
            continue
        results.append({
            "id":t.pk,
            "talep_no":t.talep_no,
            "durum":t.get_durum_display(),
            "tarih":timezone.localtime(t.olusturulma_tarihi).strftime("%d.%m.%Y %H:%M"),
            "adres":f"{t.ilce.ad} / {t.mahalle.ad} / {t.yol.ad}",
            "is":f"{t.is_turu.ad} → {t.is_alt_turu.ad}",
            "url":reverse("dashboard:talep_detay",args=[t.pk]),
        })
        if len(results) >= 5:
            break

    return JsonResponse({"ok":True,"count":len(results),"results":results})


@login_required
@require_GET
def talep_detay(request,pk):
    talep=get_object_or_404(Talep.objects.select_related(
        "ilce","mahalle","yol","is_turu","is_alt_turu",
        "sorumlu_koordinator__kullanici","sorumlu_saha__kullanici"
    ),pk=pk)
    if not talep_erisim_var_mi(request.user,talep):
        raise PermissionDenied
    # Hareket akışına kullanıcı elle mesaj yazamaz.
    return render(request,"dashboard/talep_detay.html",{
        "talep":talep,
        "surec_adimlari":_talep_surec_adimlari(talep),
    })

def _kocaeli_harita_koordinati_gecerli(lat, lng):
    """
    Operasyon haritalarında bariz hatalı / ülke dışına taşan koordinatları gösterme.
    Sınırlar Kocaeli ve yakın çevresini kapsayacak kadar geniş tutulur; amaç
    gerçek bir adresi elemek değil, bozuk koordinatın haritayı kilometrelerce
    uzaklaştırmasını önlemektir.
    """
    try:
        lat=float(lat)
        lng=float(lng)
    except (TypeError, ValueError):
        return False
    return 40.35 <= lat <= 41.35 and 29.00 <= lng <= 30.90


def _talep_harita_adresi(talep):
    parcalar=[
        talep.ilce.ad if talep.ilce_id else "",
        talep.mahalle.ad if talep.mahalle_id else "",
        talep.yol.ad if talep.yol_id else "",
        f"No: {talep.kapi_no}" if talep.kapi_no else "",
    ]
    return " / ".join(x for x in parcalar if x)


def _talep_harita_noktasi(talep):
    """
    Haritada önce talebin kendi koordinatını kullanır. Eski/bozuk bir kayıtta
    koordinat Kocaeli dışındaysa yol, mahalle veya ilçe merkezine kontrollü
    fallback yapar. Böylece tek hatalı kayıt tüm haritayı başka şehre çekmez.
    """
    adaylar=[
        (talep.lat, talep.lng, "Talep konumu", True),
        (getattr(talep.yol,"merkez_lat",None), getattr(talep.yol,"merkez_lng",None), "Yol merkezi", False) if talep.yol_id else None,
        (getattr(talep.mahalle,"merkez_lat",None), getattr(talep.mahalle,"merkez_lng",None), "Mahalle merkezi", False) if talep.mahalle_id else None,
        (getattr(talep.ilce,"merkez_lat",None), getattr(talep.ilce,"merkez_lng",None), "İlçe merkezi", False) if talep.ilce_id else None,
    ]
    for aday in adaylar:
        if not aday:
            continue
        lat,lng,kaynak,tam=aday
        if _kocaeli_harita_koordinati_gecerli(lat,lng):
            return float(lat),float(lng),kaynak,tam
    return None


@panel_required("sef","admin")
def sef(request):
    profil=operational_profile(request,"sef")
    tum_qs=kullanici_talepleri(request.user).select_related(
        "ilce","mahalle","yol","is_turu","is_alt_turu",
        "sorumlu_saha__kullanici","sorumlu_koordinator__kullanici"
    )

    # Kartlar ve yan paneller filtrelerden bağımsız operasyon özetidir.
    bekleyen_qs=tum_qs.filter(durum__in=["yeni","sefe_gonderildi"])
    sahada_qs=tum_qs.filter(durum__in=["sahaya_atandi","kabul_edildi","yolda","yerinde","islemde"])
    onay_qs=tum_qs.filter(durum="onay_bekliyor")
    tamam_qs=tum_qs.filter(durum="tamamlandi")
    acil_qs=tum_qs.filter(oncelik="acil").exclude(durum__in=["tamamlandi","iptal"])

    if request.user.is_superuser:
        saha_personelleri=PersonelProfili.objects.filter(
            aktif=True,rol__panel_tipi="saha"
        ).select_related("kullanici","rol").prefetch_related("yetkili_ilceler")
        yetkili_ilceler=Ilce.objects.filter(aktif=True)
        uzmanliklar=IsTuru.objects.filter(aktif=True)
    else:
        yetkili_ilceler=profil.yetkili_ilceler.filter(aktif=True).order_by("ad")
        uzmanliklar=profil.uzmanlik_is_turleri.filter(aktif=True).order_by("ad")
        hedef_rol_ids=RolAtamaKurali.objects.filter(
            kaynak_rol=profil.rol,aktif=True
        ).values_list("hedef_rol_id",flat=True)
        saha_personelleri=PersonelProfili.objects.filter(
            aktif=True,
            rol__panel_tipi="saha",
            rol_id__in=hedef_rol_ids,
            yetkili_ilceler__in=yetkili_ilceler,
        ).select_related("kullanici","rol").prefetch_related("yetkili_ilceler").distinct()

    # "Müsait" bayrağı yeni görev kabul tercihidir; aktif operasyonda olmayı
    # tek başına anlatmaz. Şef ekranında iki kavramı karıştırmamak için gerçek
    # operasyon durumu ayrıca hesaplanır.
    for saha_profili in saha_personelleri:
        saha_profili.acik_is_sayisi=saha_profili.saha_talepleri.exclude(
            durum__in=["tamamlandi","iptal"]
        ).count()
        saha_profili.aktif_operasyon_sayisi=saha_profili.saha_talepleri.filter(
            durum__in=["kabul_edildi","yolda","yerinde","islemde"]
        ).count()
        saha_profili.aktif_operasyon_var=saha_profili.aktif_operasyon_sayisi > 0

    son_geri_bildirimler=GeriBildirim.objects.filter(
        talep__in=tum_qs
    ).select_related("talep","kullanici").order_by("-tarih")[:6]

    qs=tum_qs
    filtre=TalepFiltreForm(request.GET)
    if filtre.is_valid():
        c=filtre.cleaned_data
        if c["durum"]: qs=qs.filter(durum=c["durum"])
        if c["oncelik"]: qs=qs.filter(oncelik=c["oncelik"])
        if c["ilce"]: qs=qs.filter(ilce=c["ilce"])
        if c["is_turu"]: qs=qs.filter(is_turu=c["is_turu"])
        if c["ara"]:
            qs=qs.filter(
                Q(talep_no__icontains=c["ara"])|
                Q(vatandas_ad__icontains=c["ara"])|
                Q(vatandas_soyad__icontains=c["ara"])|
                Q(telefon__icontains=c["ara"])
            )

    atama_bekleyen_liste=_sla_ekle(list(
        bekleyen_qs.filter(sorumlu_saha__isnull=True).order_by("olusturulma_tarihi")[:100]
    ))
    atama_bekleyen_liste.sort(key=lambda t:t.sla["kalan_dakika"])
    sla_geciken=[t for t in atama_bekleyen_liste if t.sla["durum"]=="asildi"]
    sla_yaklasan=[t for t in atama_bekleyen_liste if t.sla["durum"]=="yaklasiyor"]

    # V37 — Şef haritası yalnız AKTİF işleri gösterir. Tamamlanan kayıtların
    # harita sınırlarını büyütmesi ve aktif işlerin kaybolması engellenir.
    # Her pin Talep.lat/lng değerinin kendisidir; pin ile liste kartı birebir bağlıdır.
    aktif_harita_qs=tum_qs.exclude(
        durum__in=["tamamlandi","iptal"]
    ).select_related("ilce","mahalle","yol","is_turu","is_alt_turu")
    harita_json=[]
    for t in aktif_harita_qs[:300]:
        nokta=_talep_harita_noktasi(t)
        if not nokta:
            continue
        lat,lng,konum_kaynagi,tam_konum=nokta
        if t.durum in ("yeni","sefe_gonderildi"):
            harita_durum="bekleyen"
        elif t.durum=="onay_bekliyor":
            harita_durum="onay"
        else:
            harita_durum="sahada"
        harita_json.append({
            "id":t.pk,
            "no":t.talep_no,
            "lat":lat,
            "lng":lng,
            "durum":harita_durum,
            "durum_label":t.get_durum_display(),
            "ilce":t.ilce.ad,
            "mahalle":t.mahalle.ad,
            "adres":_talep_harita_adresi(t),
            "tur":t.is_turu.ad,
            "alt_tur":t.is_alt_turu.ad,
            "oncelik":t.get_oncelik_display(),
            "konum_kaynagi":konum_kaynagi,
            "tam_konum":tam_konum,
            "url":reverse("dashboard:talep_detay",args=[t.pk]),
        })

    harita_idleri={x["id"] for x in harita_json}
    for t in atama_bekleyen_liste:
        t.haritada_var=t.pk in harita_idleri

    from abonelik.models import AmbarYetkisi
    ambar_yetkisi = AmbarYetkisi.objects.select_related("ilce", "ambar").filter(personel=profil, aktif=True).first()

    return render(request,"dashboard/sef.html",{
        "talepler":qs[:200],
        "filtre":filtre,
        "bekleyen":bekleyen_qs.count(),
        "sahada":sahada_qs.count(),
        "onay_bekleyen":onay_qs.count(),
        "tamam":tamam_qs.count(),
        "acil":acil_qs.count(),
        "atama_bekleyen":atama_bekleyen_liste,
        "sla_geciken":sla_geciken,
        "sla_yaklasan":sla_yaklasan,
        "sla_geciken_sayisi":len(sla_geciken),
        "sla_yaklasan_sayisi":len(sla_yaklasan),
        "sla_kurallari":SLA_ATAMA_DAKIKA,
        "canli_operasyonlar":(
            sahada_qs | onay_qs
        ).distinct().order_by("-guncellenme_tarihi")[:10],
        "onay_bekleyen_talepler":onay_qs.select_related(
            "sorumlu_saha__kullanici","sorumlu_saha__rol"
        ).order_by("-saha_tamam_bildirim_tarihi")[:20],
        "saha_personelleri":saha_personelleri,
        "son_geri_bildirimler":son_geri_bildirimler,
        "yetkili_ilceler":yetkili_ilceler,
        "uzmanliklar":uzmanliklar,
        "harita_json":harita_json,
        "ambar_yetkisi":ambar_yetkisi,
    })

@panel_required("sef","admin")
def talep_ata(request,pk):
    p=operational_profile(request,"sef")
    talep=get_object_or_404(
        Talep.objects.select_related("ilce","mahalle","is_turu","is_alt_turu"),
        pk=pk
    )

    if not talep_erisim_var_mi(request.user,talep):
        raise PermissionDenied
    if talep.sorumlu_koordinator_id and talep.sorumlu_koordinator_id!=p.id:
        raise PermissionDenied("Bu talep başka bir koordinatörün sorumluluğunda.")
    if talep.sorumlu_saha_id or talep.durum not in ["yeni","sefe_gonderildi"]:
        messages.error(request,"Bu talep zaten sahaya atanmış veya operasyon aşamasına geçmiş.")
        return redirect("dashboard:sef")

    uygun=uygun_saha_personelleri(p,talep)
    bolge_saha=bolge_saha_personelleri(p,talep)

    # V35 — Şef manuel seçim yaparken her ekibin mevcut iş yükünü görür.
    # Bu bilgi seçim önerisi/puanlama değildir; yalnız operasyonel görünürlüktür.
    is_yuku={}
    for saha_profili in bolge_saha:
        is_yuku[saha_profili.pk]=saha_profili.saha_talepleri.exclude(
            durum__in=["tamamlandi","iptal"]
        ).count()
        saha_profili.acik_is_sayisi=is_yuku[saha_profili.pk]
        saha_profili.aktif_operasyon_sayisi=saha_profili.saha_talepleri.filter(
            durum__in=["kabul_edildi","yolda","yerinde","islemde"]
        ).count()
        saha_profili.aktif_operasyon_var=saha_profili.aktif_operasyon_sayisi > 0

    for saha_profili in uygun:
        saha_profili.acik_is_sayisi=is_yuku.get(
            saha_profili.pk,
            saha_profili.saha_talepleri.exclude(durum__in=["tamamlandi","iptal"]).count()
        )
        saha_profili.aktif_operasyon_sayisi=saha_profili.saha_talepleri.filter(
            durum__in=["kabul_edildi","yolda","yerinde","islemde"]
        ).count()
        saha_profili.aktif_operasyon_var=saha_profili.aktif_operasyon_sayisi > 0

    if request.method=="POST":
        saha_id=request.POST.get("saha")
        saha=uygun.filter(pk=saha_id).first()
        if not saha:
            messages.error(request,"Seçilen personel ilçe/uzmanlık/müsaitlik kurallarına uygun değil.")
            return redirect("dashboard:talep_ata",pk=pk)
        is_emri=_saha_atamasini_yap(request,talep,saha,p)
        messages.success(request,f"Atama tamamlandı. {is_emri.is_emri_no} oluşturuldu.")
        return redirect("dashboard:sef")

    return render(request,"dashboard/talep_ata.html",{
        "talep":talep,
        "uygun":uygun,
        "bolge_saha":bolge_saha,
        "atan_kisi":p,
    })



def _saha_web_context(p, request, abone_web=False):
    qs=kullanici_talepleri(request.user).select_related(
        "ilce","mahalle","yol","is_turu","is_alt_turu",
        "sorumlu_koordinator__kullanici","is_emri"
    ).prefetch_related("is_emri__fotograflar")

    aktif_liste=list(qs.exclude(durum__in=["tamamlandi","iptal"]).order_by("-guncellenme_tarihi"))
    for t in aktif_liste:
        if hasattr(t,"is_emri") and t.is_emri:
            slotlar,yuklenen,zorunlu=_is_emri_foto_durumu(t.is_emri)
            t.foto_slotlari=slotlar
            t.foto_yuklenen=yuklenen
            t.foto_zorunlu=zorunlu
        else:
            t.foto_slotlari=[]; t.foto_yuklenen=0; t.foto_zorunlu=0

    tamam_liste=list(qs.filter(durum="tamamlandi").order_by("-tamamlanma_tarihi")[:8])
    return {
        "talepler":aktif_liste,
        "tamamlananlar":tamam_liste,
        "profil":p,
        "abone_web":abone_web,
        "aktif":len(aktif_liste),
        "tamam":qs.filter(durum="tamamlandi").count(),
        "acil":sum(1 for t in aktif_liste if t.oncelik=="acil"),
    }


@panel_required("saha")
def abone_islemleri(request):
    """V56: Abone Personeli ayrı Web/PC rolü olarak Sicil + Sözleşme + Sayaç merkezine gider."""
    p=get_profile(request.user)
    if not p or not p.aktif or p.rol.panel_tipi != "abone" or not _web_only_abone_role(p.rol):
        raise PermissionDenied("Bu ekran yalnız Abone Personeline aittir.")
    return redirect("abonelik:izleme")


@panel_required("saha")
def abone_gorevleri(request):
    """Eski atanmış abonelik/sayaç iş emirleri korunur; yeni izleme ekranından ayrı sekmede açılır."""
    p=operational_profile(request,"saha")
    if not _web_only_abone_role(p.rol):
        raise PermissionDenied("Bu ekran yalnız abonelik/sayaç web rolüne aittir.")
    return render(request,"dashboard/saha.html",_saha_web_context(p,request,abone_web=True))


@panel_required("saha","admin")
def saha(request):
    p=operational_profile(request,"saha")
    if _web_only_abone_role(p.rol):
        return redirect("dashboard:abone_islemleri")
    if p.rol.calisma_kanali=="mobil":
        return redirect("dashboard:mobil_saha")

    qs=kullanici_talepleri(request.user).select_related(
        "ilce","mahalle","yol","is_turu","is_alt_turu",
        "sorumlu_koordinator__kullanici","is_emri"
    ).prefetch_related("is_emri__fotograflar")

    aktif_liste=list(qs.exclude(durum__in=["tamamlandi","iptal"]).order_by("-guncellenme_tarihi"))
    for t in aktif_liste:
        if hasattr(t,"is_emri") and t.is_emri:
            slotlar,yuklenen,zorunlu=_is_emri_foto_durumu(t.is_emri)
            t.foto_slotlari=slotlar
            t.foto_yuklenen=yuklenen
            t.foto_zorunlu=zorunlu
        else:
            t.foto_slotlari=[];t.foto_yuklenen=0;t.foto_zorunlu=0

    tamam_liste=list(qs.filter(durum="tamamlandi").order_by("-tamamlanma_tarihi")[:8])

    return render(request,"dashboard/saha.html",{
        "talepler":aktif_liste,
        "tamamlananlar":tamam_liste,
        "profil":p,
        "abone_web":_web_only_abone_role(p.rol),
        "aktif":len(aktif_liste),
        "tamam":qs.filter(durum="tamamlandi").count(),
        "acil":sum(1 for t in aktif_liste if t.oncelik=="acil"),
    })


@panel_required("saha","admin")
@require_POST
def durum_guncelle(request,pk,yeni_durum):
    p=operational_profile(request,"saha")
    mobil=request.POST.get("next")=="mobil"
    if _web_only_abone_role(p.rol):
        mobil=False
        hedef="dashboard:abone_gorevleri"
    else:
        hedef="dashboard:mobil_saha" if mobil else "dashboard:saha"
    talep=get_object_or_404(Talep,pk=pk)

    if talep.sorumlu_saha_id!=p.id:
        raise PermissionDenied

    gecisler={
        "sahaya_atandi":("kabul_edildi","Talep saha ekibi tarafından kabul edildi."),
        "kabul_edildi":("yolda","Saha ekibi arıza noktasına doğru yola çıktı."),
        "yolda":("yerinde","Saha ekibi arıza adresine ulaştı ve yerinde incelemeye başladı."),
        "yerinde":("islemde","Yerinde inceleme tamamlandı; arızaya müdahale başladı."),
        "islemde":("onay_bekliyor","Saha ekibi işi tamamladığını bildirdi; talep şef onayına gönderildi."),
    }
    beklenen=gecisler.get(talep.durum)
    if not beklenen or beklenen[0]!=yeni_durum:
        messages.error(request,"İş akışı sırası geçersiz. Önce mevcut aşamanın sıradaki işlemini tamamlayın.")
        return redirect(hedef)

    # V49 — Bir saha ekibine birden fazla iş emri atanabilir ve ekip bu işler
    # arasında serbestçe geçiş yapabilir. Başka bir iş kabul/yol/yerinde/islemde
    # aşamasında olsa bile yeni atanmış işin kabul edilmesi engellenmez.

    eski=talep.durum
    emri=IsEmri.objects.filter(talep=talep).first()
    if not emri:
        emri=_is_emri_esitle(talep,request.user)

    # V46 — İş kabulü ve yola çıkış her yerden yapılabilir.
    # Sahada fiziksel bulunmayı gerektiren adımlar anlık GPS ile doğrulanır.
    gps_mesafe_gecis=None
    gps_aksiyon=None
    if mobil and eski=="yolda" and yeni_durum=="yerinde":
        gps_aksiyon="Adrese Ulaştım"
    elif mobil and eski=="yerinde" and yeni_durum=="islemde":
        gps_aksiyon="Müdahaleye Başla"

    if gps_aksiyon:
        gps_ok,gps_mesafe_gecis,gps_hata=_mobil_saha_gps_dogrula(request,talep,emri,gps_aksiyon)
        if not gps_ok:
            messages.error(request,gps_hata)
            return redirect(hedef)

    # V42 — Şef onayına geçmeden önce iş alt türünün zorunlu görsel slotları tamamlanmalıdır.
    if eski=="islemde" and yeni_durum=="onay_bekliyor":
        slotlar,yuklenen,zorunlu=_is_emri_foto_durumu(emri)
        eksikler=[x["etiket"] for x in slotlar if not x["tamam"]]
        if eksikler:
            messages.error(
                request,
                f"Bu iş alt türü için {zorunlu} zorunlu fotoğraf gerekiyor. "
                f"Eksik: {', '.join(eksikler)}."
            )
            return redirect(hedef)

    # İş bitiminde sonuç notu zorunludur.
    if eski=="islemde" and yeni_durum=="onay_bekliyor":
        sonuc_notu=(request.POST.get("sonuc_notu") or "").strip()
        if not sonuc_notu:
            messages.error(request,"Şef onayına göndermeden önce saha sonuç notunu yazın.")
            return redirect(hedef)
        if len(sonuc_notu)>1000:
            messages.error(request,"Saha sonuç notu en fazla 1000 karakter olabilir.")
            return redirect(hedef)

        talep.saha_sonuc_notu=sonuc_notu
        talep.saha_tamam_bildirim_tarihi=timezone.now()
        talep.sef_onaylayan=None
        talep.sef_onay_tarihi=None

    talep.durum=yeni_durum
    talep.save()
    emri=_is_emri_esitle(talep,request.user)

    hareket_mesaji=beklenen[1]
    ekler=[]
    if gps_aksiyon and emri.gps_dogrulama_tarihi:
        if gps_mesafe_gecis is not None:
            ekler.append(f"{gps_aksiyon} GPS mesafesi: {gps_mesafe_gecis} m")
        ekler.append("GPS doğrulandı" if emri.gps_dogrulandi else "GPS koordinatı kaydedildi")
    if eski=="islemde" and yeni_durum=="onay_bekliyor":
        hareket_mesaji=(
            "Saha ekibi işi tamamladığını bildirdi ve şef onayına gönderdi. "
            f"Saha sonuç notu: {talep.saha_sonuc_notu}"
        )
        ekler.append(f"{yuklenen}/{zorunlu} zorunlu iş emri görseli tamamlandı")
    if ekler:
        hareket_mesaji += " • " + " • ".join(ekler)

    akisa_yaz(talep,request.user,hareket_mesaji,yeni_durum,request,True,"FIELD_STATUS")
    log_yaz(
        request,"FIELD_STATUS_CHANGE",hareket_mesaji,talep,"IsEmri",emri.pk,eski,yeni_durum
    )
    messages.success(request,beklenen[1])
    return redirect(hedef)


@panel_required("saha","admin")
@require_POST
def saha_notu(request,pk):
    """
    Manuel geri bildirim değildir.
    Saha personelinin operasyonel açıklamasını kaydeder; sistem bu işlemi
    otomatik hareket ve teknik log olarak yayınlar.
    """
    p=operational_profile(request,"saha")
    talep=get_object_or_404(Talep,pk=pk)

    if talep.sorumlu_saha_id!=p.id:
        raise PermissionDenied

    if talep.durum in ["onay_bekliyor","tamamlandi","iptal"]:
        messages.error(request,"Şef onayı bekleyen veya kapanmış talebe yeni saha açıklaması eklenemez.")
        return redirect("dashboard:saha")

    not_metni=(request.POST.get("not") or "").strip()
    if not not_metni:
        messages.error(request,"Saha açıklaması boş bırakılamaz.")
        return redirect("dashboard:saha")
    if len(not_metni)>700:
        messages.error(request,"Saha açıklaması en fazla 700 karakter olabilir.")
        return redirect("dashboard:saha")

    mesaj=f"Saha açıklaması kaydedildi: {not_metni}"
    akisa_yaz(
        talep,
        request.user,
        mesaj,
        talep.durum,
        request,
        True,
        "FIELD_NOTE",
    )
    log_yaz(
        request,
        "FIELD_NOTE",
        mesaj,
        talep,
        "Talep",
        talep.pk,
    )
    messages.success(request,"Saha açıklaması kaydedildi ve hareket akışına işlendi.")
    return redirect("dashboard:saha")



@panel_required("sef","admin")
@require_POST
def sef_tamamla_onayla(request,pk):
    p=operational_profile(request,"sef")
    talep=get_object_or_404(Talep,pk=pk)

    if not talep_erisim_var_mi(request.user,talep):
        raise PermissionDenied

    if talep.sorumlu_koordinator_id and talep.sorumlu_koordinator_id!=p.id:
        raise PermissionDenied("Bu talep başka bir koordinatörün sorumluluğunda.")

    if talep.durum!="onay_bekliyor":
        messages.error(request,"Bu talep şu anda şef onayı beklemiyor.")
        return redirect("dashboard:sef")

    eski=talep.durum
    talep.durum="tamamlandi"
    talep.sef_onaylayan=request.user
    talep.sef_onay_tarihi=timezone.now()

    # Teknik operasyon tamamlandı; şimdi 185'in vatandaşa dönüş işi başlar.
    talep.vatandas_bildirim_durumu="bekliyor"
    talep.vatandas_bildirim_tarihi=None
    talep.vatandas_bildirim_yapan=None
    talep.save()
    _is_emri_esitle(talep,request.user)

    mesaj=(
        "Saha tarafından tamamlandı olarak bildirilen iş şef/koordinatör tarafından "
        "kontrol edilerek onaylandı. Operasyon tamamlandı ve talep 185 Vatandaş "
        "Bilgilendirme kuyruğuna gönderildi."
    )
    akisa_yaz(
        talep,request.user,mesaj,"tamamlandi",request,True,"CHIEF_APPROVE"
    )
    log_yaz(
        request,"CHIEF_APPROVE",mesaj,talep,"Talep",talep.pk,eski,"tamamlandi"
    )
    messages.success(
        request,
        f"{talep.talep_no} onaylandı. Operasyon Tamamlandı; 185 vatandaş bilgilendirmesi bekliyor."
    )
    return redirect("dashboard:sef")


@panel_required("sef","admin")
@require_POST
def sef_tamamla_iade(request,pk):
    p=operational_profile(request,"sef")
    talep=get_object_or_404(Talep,pk=pk)

    if not talep_erisim_var_mi(request.user,talep):
        raise PermissionDenied

    if talep.sorumlu_koordinator_id and talep.sorumlu_koordinator_id!=p.id:
        raise PermissionDenied("Bu talep başka bir koordinatörün sorumluluğunda.")

    if talep.durum!="onay_bekliyor":
        messages.error(request,"Bu talep şu anda şef onayı beklemiyor.")
        return redirect("dashboard:sef")

    gerekce=(request.POST.get("gerekce") or "").strip()
    if not gerekce:
        messages.error(request,"Sahaya geri gönderirken gerekçe yazılması zorunludur.")
        return redirect("dashboard:sef")
    if len(gerekce)>700:
        messages.error(request,"İade gerekçesi en fazla 700 karakter olabilir.")
        return redirect("dashboard:sef")

    eski=talep.durum
    ayni_talep_no=talep.talep_no

    # V37 — Aynı kayıt yeniden sahaya döner; yeni Talep oluşturulmaz ve talep_no değişmez.
    # Geri gönderilen işi doğrudan "Müdahale Ediliyor" durumuna almak, ekip bu sırada
    # başka bir işi yürütüyorsa tek-aktif-operasyon kuralını bozuyordu. Bu nedenle iş
    # aynı saha hesabının KUYRUĞUNA geri düşer ve ekip yeniden kabul ederek akışı başlatır.
    talep.durum="sahaya_atandi"
    talep.sef_onaylayan=None
    talep.sef_onay_tarihi=None

    # Önceki saha sonuç notu hareket akışında zaten saklıdır.
    # Yeni müdahale turu için sonuç alanını temizleyip taze rapor bekleriz.
    talep.saha_sonuc_notu=""
    talep.saha_tamam_bildirim_tarihi=None
    talep.save()
    emri=_is_emri_esitle(talep,request.user)
    if emri:
        # Önceki turun "saha tamam" bilgisi aktif kuyruğa geri dönen işte
        # güncel tamamlanma gibi görünmesin. Tekrar geçmişi hareket/loglarda korunur.
        emri.saha_tamam_tarihi=None
        emri.sef_onay_tarihi=None
        emri.sonuc_notu=""
        emri.save(update_fields=["saha_tamam_tarihi","sef_onay_tarihi","sonuc_notu","guncellenme_tarihi"])
    if talep.sorumlu_saha_id and emri:
        _mobil_bildirim_olustur(
            talep.sorumlu_saha.kullanici,
            emri,
            "geri_gonderildi",
            f"İş Emri Geri Gönderildi • {emri.is_emri_no}",
            f"Şef yeniden müdahale istedi. Gerekçe: {gerekce}",
        )

    if talep.talep_no != ayni_talep_no:
        raise RuntimeError("Talep numarası yeniden saha gönderiminde değiştirilemez.")

    mesaj=(
        f"{ayni_talep_no} numaralı aynı talep, yeni kayıt oluşturulmadan "
        "şef/koordinatör tarafından yeniden sahaya gönderildi. "
        f"Gerekçe: {gerekce}"
    )
    akisa_yaz(
        talep,request.user,mesaj,"sahaya_atandi",request,True,"CHIEF_RETURN"
    )
    log_yaz(
        request,"CHIEF_RETURN",mesaj,talep,"Talep",talep.pk,eski,"sahaya_atandi"
    )
    messages.warning(
        request,
        f"{talep.talep_no} numarası korunarak aynı talep saha ekibinin iş kuyruğuna geri gönderildi."
    )
    return redirect("dashboard:sef")


@panel_required("saha","admin")
@require_POST
def musaitlik(request):
    p=operational_profile(request,"saha")
    p.musait=not p.musait
    p.save(update_fields=["musait"])
    log_yaz(
        request,
        "AVAILABILITY",
        f"Yeni görev kabul durumu: {'Açık' if p.musait else 'Kapalı'}",
        varlik_turu="Personel",
        varlik_id=p.pk
    )
    messages.success(
        request,
        f"Yeni görev kabulü {'açıldı' if p.musait else 'kapatıldı'}."
    )
    if _web_only_abone_role(p.rol):
        return redirect("dashboard:abone_gorevleri")
    return redirect("dashboard:saha")



@panel_required("185","admin")
@require_POST
def vatandas_arama_baslat(request,pk):
    operational_profile(request,"185")
    talep=get_object_or_404(Talep,pk=pk)

    if talep.durum!="tamamlandi" or talep.vatandas_bildirim_durumu=="bilgilendirildi":
        return JsonResponse({
            "ok":False,
            "message":"Bu talep vatandaş arama kuyruğunda değil."
        },status=400)

    son_baslatma=IslemLogu.objects.filter(
        talep=talep,islem="CITIZEN_CALL_START"
    ).select_related("kullanici").order_by("-tarih").first()
    son_sonuc=IslemLogu.objects.filter(
        talep=talep,islem="CITIZEN_CALLBACK"
    ).order_by("-tarih").first()

    # Sonuçlanmamış bir arama başka operatörde açıksa ikinci personel aynı vatandaşı aramasın.
    if son_baslatma and (not son_sonuc or son_baslatma.tarih > son_sonuc.tarih):
        if son_baslatma.kullanici_id != request.user.id:
            kullanici=(
                son_baslatma.kullanici.get_full_name()
                or son_baslatma.kullanici.username
                if son_baslatma.kullanici else "başka bir 185 personeli"
            )
            return JsonResponse({
                "ok":False,
                "locked":True,
                "message":f"Bu vatandaş için arama oturumu {kullanici} tarafından açık."
            },status=409)

        return JsonResponse({
            "ok":True,
            "already_started":True,
            "started_at":timezone.localtime(son_baslatma.tarih).strftime("%H:%M:%S"),
            "message":"Açık arama oturumunuz devam ediyor."
        })

    aciklama=(
        f"185 Çağrı Merkezi vatandaş arama oturumunu başlattı. "
        f"Talep: {talep.talep_no}. Telefon: {talep.telefon}. "
        "Görüşme sonucu henüz girilmedi."
    )
    log_yaz(
        request,
        "CITIZEN_CALL_START",
        aciklama,
        talep,
        "Talep",
        talep.pk,
        "",
        "arama_baslatildi",
    )
    akisa_yaz(
        talep,request.user,
        "185 Çağrı Merkezi vatandaş geri dönüş aramasını başlattı.",
        talep.durum,None,True,"CITIZEN_CALL_START"
    )

    return JsonResponse({
        "ok":True,
        "already_started":False,
        "started_at":timezone.localtime(timezone.now()).strftime("%H:%M:%S"),
        "message":"Arama oturumu kullanıcı, IP ve saat bilgisiyle kaydedildi."
    })


@panel_required("185","admin")
@require_POST
def vatandas_bildirim_kaydet(request,pk):
    operational_profile(request,"185")
    talep=get_object_or_404(Talep,pk=pk)

    if talep.durum!="tamamlandi":
        messages.error(
            request,
            "Geri bildirim yalnız şef tarafından onaylanmış tamamlanan işler için kaydedilebilir."
        )
        return redirect(f"{reverse('dashboard:cagri_185')}#vatandas-bilgilendirme")

    if talep.vatandas_bildirim_durumu=="bilgilendirildi":
        messages.info(
            request,
            "Bu talebin vatandaş geri bildirim süreci daha önce tamamlanmış."
        )
        return redirect(f"{reverse('dashboard:cagri_185')}#vatandas-bilgilendirme")

    sonuc=(request.POST.get("sonuc") or "").strip()
    memnuniyet=(request.POST.get("memnuniyet") or "").strip()
    sorun_cozuldu=(request.POST.get("sorun_cozuldu") or "").strip()
    hizmet_hizi=(request.POST.get("hizmet_hizi") or "").strip()
    bilgilendirme=(request.POST.get("bilgilendirme") or "").strip()
    personel_iletisimi=(request.POST.get("personel_iletisimi") or "").strip()
    genel_puan=(request.POST.get("genel_puan") or "").strip()
    islem_suresi_raw=(request.POST.get("islem_suresi") or "").strip()
    not_metni=(request.POST.get("arama_notu") or "").strip()

    allowed={"bilgilendirildi","ulasilamadi","tekrar_aranacak"}
    memnuniyet_allowed={"iyi":"İyi","normal":"Normal","kotu":"Kötü"}
    sorun_allowed={"evet":"Evet","kismen":"Kısmen","hayir":"Hayır"}
    hiz_allowed={"hizli":"Hızlı","normal":"Normal","yavas":"Yavaş"}
    bilgi_allowed={"yeterli":"Yeterli","kismen":"Kısmen","yetersiz":"Yetersiz"}
    iletisim_allowed={"iyi":"İyi","normal":"Normal","kotu":"Kötü"}
    puan_allowed={"1","2","3","4","5"}

    if sonuc not in allowed:
        messages.error(request,"Görüşme sonucunu seçin.")
        return redirect(f"{reverse('dashboard:cagri_185')}#vatandas-bilgilendirme")

    if len(not_metni)>300:
        messages.error(request,"Görüşme notu en fazla 300 karakter olabilir.")
        return redirect(f"{reverse('dashboard:cagri_185')}#vatandas-bilgilendirme")

    islem_suresi=None

    # Vatandaşla görüşme sağlandıysa değerlendirme alanları zorunludur.
    if sonuc=="bilgilendirildi":
        if memnuniyet not in memnuniyet_allowed:
            messages.error(
                request,
                "Vatandaşın genel memnuniyetini seçin: İyi, Normal veya Kötü."
            )
            return redirect(f"{reverse('dashboard:cagri_185')}#vatandas-bilgilendirme")

        try:
            islem_suresi=int(islem_suresi_raw)
        except (TypeError,ValueError):
            islem_suresi=None

        if not islem_suresi or islem_suresi<1 or islem_suresi>1440:
            messages.error(
                request,
                "İşlem süresini dakika olarak 1 ile 1440 arasında girin."
            )
            return redirect(f"{reverse('dashboard:cagri_185')}#vatandas-bilgilendirme")

        # V19: Alttaki vatandaş memnuniyet anketi tamamen opsiyoneldir.
        # Bir soru cevaplandıysa yalnızca o cevabın geçerli seçeneklerden olması kontrol edilir.
        if sorun_cozuldu and sorun_cozuldu not in sorun_allowed:
            messages.error(request,"Anketteki 'Sorun Çözüldü mü?' cevabı geçersiz.")
            return redirect(f"{reverse('dashboard:cagri_185')}#vatandas-bilgilendirme")
        if hizmet_hizi and hizmet_hizi not in hiz_allowed:
            messages.error(request,"Anketteki hizmet hızı cevabı geçersiz.")
            return redirect(f"{reverse('dashboard:cagri_185')}#vatandas-bilgilendirme")
        if bilgilendirme and bilgilendirme not in bilgi_allowed:
            messages.error(request,"Anketteki bilgilendirme cevabı geçersiz.")
            return redirect(f"{reverse('dashboard:cagri_185')}#vatandas-bilgilendirme")
        if personel_iletisimi and personel_iletisimi not in iletisim_allowed:
            messages.error(request,"Anketteki personel iletişimi cevabı geçersiz.")
            return redirect(f"{reverse('dashboard:cagri_185')}#vatandas-bilgilendirme")
        if genel_puan and genel_puan not in puan_allowed:
            messages.error(request,"Genel hizmet puanı 1 ile 5 arasında olmalıdır.")
            return redirect(f"{reverse('dashboard:cagri_185')}#vatandas-bilgilendirme")
    else:
        # Ulaşılamadı / tekrar aranacak için değerlendirme alanları kullanılmaz.
        memnuniyet=""
        sorun_cozuldu=""
        hizmet_hizi=""
        bilgilendirme=""
        personel_iletisimi=""
        genel_puan=""
        islem_suresi=None

    # Mevcut VatandasAramaKaydi tablosunu değiştirmeden görüşme detayını
    # yapılandırılmış ve okunabilir biçimde not alanında sakla.
    detay_parcalari=[]
    if sonuc=="bilgilendirildi":
        detay_parcalari.append(
            f"Genel Memnuniyet: {memnuniyet_allowed[memnuniyet]}"
        )
        detay_parcalari.append(
            f"İşlem Süresi: {islem_suresi} dk"
        )
        # Opsiyonel anket: yalnız cevaplanan soruları sakla.
        if sorun_cozuldu:
            detay_parcalari.append(f"Anket / Sorun Çözüldü mü: {sorun_allowed[sorun_cozuldu]}")
        if hizmet_hizi:
            detay_parcalari.append(f"Anket / Hizmet Hızı: {hiz_allowed[hizmet_hizi]}")
        if bilgilendirme:
            detay_parcalari.append(f"Anket / Bilgilendirme: {bilgi_allowed[bilgilendirme]}")
        if personel_iletisimi:
            detay_parcalari.append(f"Anket / Personel İletişimi: {iletisim_allowed[personel_iletisimi]}")
        if genel_puan:
            detay_parcalari.append(f"Anket / Genel Hizmet Puanı: {genel_puan}/5")
    if not_metni:
        detay_parcalari.append(f"Görüşme Notu: {not_metni}")
    yapilandirilmis_not=" | ".join(detay_parcalari)

    kayit=VatandasAramaKaydi.objects.create(
        talep=talep,
        kullanici=request.user,
        sonuc=sonuc,
        not_metni=yapilandirilmis_not,
    )

    if sonuc=="bilgilendirildi":
        talep.vatandas_bildirim_durumu="bilgilendirildi"
        talep.vatandas_bildirim_tarihi=timezone.now()
        talep.vatandas_bildirim_yapan=request.user
        hareket=(
            "185 Çağrı Merkezi vatandaşla geri bildirim görüşmesini tamamladı. "
            f"Genel memnuniyet: {memnuniyet_allowed[memnuniyet]}. "
            f"İşlem süresi: {islem_suresi} dakika."
        )
        anket_ozeti=[]
        if sorun_cozuldu:
            anket_ozeti.append(f"sorun çözümü {sorun_allowed[sorun_cozuldu]}")
        if hizmet_hizi:
            anket_ozeti.append(f"hizmet hızı {hiz_allowed[hizmet_hizi]}")
        if bilgilendirme:
            anket_ozeti.append(f"bilgilendirme {bilgi_allowed[bilgilendirme]}")
        if personel_iletisimi:
            anket_ozeti.append(f"personel iletişimi {iletisim_allowed[personel_iletisimi]}")
        if genel_puan:
            anket_ozeti.append(f"genel puan {genel_puan}/5")
        if anket_ozeti:
            hareket += " Opsiyonel anket: " + ", ".join(anket_ozeti) + "."
        if not_metni:
            hareket+=f" Görüşme notu: {not_metni}"
        flash=f"{talep.talep_no}: geri bildirim kaydedildi ve vatandaş kapanışı tamamlandı."

    elif sonuc=="ulasilamadi":
        talep.vatandas_bildirim_durumu="tekrar_aranacak"
        hareket=(
            "185 Çağrı Merkezi vatandaşa ulaşamadı. "
            "Kayıt tekrar aranacaklar kuyruğunda tutuluyor."
        )
        if not_metni:
            hareket+=f" Görüşme notu: {not_metni}"
        flash=f"{talep.talep_no}: vatandaşa ulaşılamadı; kuyrukta tutuluyor."

    else:
        talep.vatandas_bildirim_durumu="tekrar_aranacak"
        hareket=(
            "185 Çağrı Merkezi geri dönüşü tekrar aranacak olarak kaydetti."
        )
        if not_metni:
            hareket+=f" Görüşme notu: {not_metni}"
        flash=f"{talep.talep_no}: tekrar aranacak olarak kaydedildi."

    talep.save(update_fields=[
        "vatandas_bildirim_durumu",
        "vatandas_bildirim_tarihi",
        "vatandas_bildirim_yapan",
        "guncellenme_tarihi",
    ])

    akisa_yaz(
        talep,request.user,hareket,talep.durum,request,True,"CITIZEN_CALLBACK"
    )
    log_yaz(
        request,
        "CITIZEN_CALLBACK",
        hareket,
        talep,
        "VatandasAramaKaydi",
        kayit.pk,
        "",
        sonuc,
    )

    messages.success(request,flash)
    return redirect(f"{reverse('dashboard:cagri_185')}#vatandas-bilgilendirme")


@panel_required("admin")
def sistem(request):
    qs=Talep.objects.select_related(
        "ilce","mahalle","yol","is_turu","is_alt_turu"
    )
    by={x["durum"]:x["adet"] for x in qs.values("durum").annotate(adet=Count("id"))}

    map_points=[]
    for t in qs.exclude(lat__isnull=True).exclude(lng__isnull=True)[:500]:
        map_points.append({
            "id":t.id,
            "no":t.talep_no,
            "lat":float(t.lat),
            "lng":float(t.lng),
            "durum":t.durum,
            "durum_label":t.get_durum_display(),
            "oncelik":t.oncelik,
            "oncelik_label":t.get_oncelik_display(),
            "ilce":t.ilce.ad,
            "mahalle":t.mahalle.ad,
            "tur":t.is_turu.ad,
            "alt_tur":t.is_alt_turu.ad,
        })

    organizasyon_ozeti=[]
    acik_saha_durumlari=["sahaya_atandi","kabul_edildi","yolda","yerinde","islemde","onay_bekliyor"]
    operasyon_durumlari=["kabul_edildi","yolda","yerinde","islemde"]
    for ilce in Ilce.objects.filter(aktif=True).order_by("ad"):
        sefler=PersonelProfili.objects.filter(
            aktif=True,
            kullanici__is_active=True,
            rol__panel_tipi="sef",
            yetkili_ilceler=ilce,
        ).select_related("kullanici","rol").distinct()

        saha_qs=PersonelProfili.objects.filter(
            aktif=True,
            kullanici__is_active=True,
            rol__panel_tipi="saha",
            yetkili_ilceler=ilce,
        ).select_related("kullanici","rol").distinct()

        aktif_saha_ids=set(
            Talep.objects.filter(
                ilce=ilce,
                durum__in=operasyon_durumlari,
                sorumlu_saha__isnull=False,
            ).values_list("sorumlu_saha_id",flat=True)
        )

        saha_list=list(saha_qs)
        organizasyon_ozeti.append({
            "ilce":ilce,
            "sefler":list(sefler[:3]),
            "sef_sayisi":sefler.count(),
            "saha_sayisi":len(saha_list),
            "musait_saha":sum(
                1 for p in saha_list
                if p.musait and p.id not in aktif_saha_ids
            ),
            "aktif_operasyon":Talep.objects.filter(
                ilce=ilce,
                durum__in=operasyon_durumlari,
            ).count(),
        })

    sla_bekleyen_qs=qs.filter(
        durum__in=["yeni","sefe_gonderildi"],
        sorumlu_saha__isnull=True,
    ).select_related(
        "ilce","mahalle","is_turu","is_alt_turu",
        "sorumlu_koordinator__kullanici",
    ).order_by("olusturulma_tarihi")[:100]

    sla_bekleyen=_sla_ekle(list(sla_bekleyen_qs))
    sla_bekleyen.sort(key=lambda t:t.sla["kalan_dakika"])
    sla_geciken=[t for t in sla_bekleyen if t.sla["durum"]=="asildi"]
    sla_yaklasan=[t for t in sla_bekleyen if t.sla["durum"]=="yaklasiyor"]

    # V56: aktif sistemde devir/ayrı ambar hesabı yok; abonelik ve ambar yapısı tek özet altında izlenir.
    from abonelik.models import AmbarSayacTalebi, SayacEnvanteri, Sozlesme, VatandasSicili
    abonelik_ozet = {
        "aktif_sicil": VatandasSicili.objects.filter(aktif=True).count(),
        "aktif_sozlesme": Sozlesme.objects.filter(aktif=True).count(),
        "takili_sayac": SayacEnvanteri.objects.filter(durum="aboneye_takili").count(),
        "merkez_stok": SayacEnvanteri.objects.filter(ambar__tur="merkez", durum="stokta").count(),
        "acik_ambar_talebi": AmbarSayacTalebi.objects.exclude(durum__in=["teslim_alindi", "reddedildi"]).count(),
        "hurda": SayacEnvanteri.objects.filter(durum="hurda").count(),
    }

    return render(request,"system/dashboard.html",{
        "toplam":qs.count(),
        "yeni":by.get("yeni",0)+by.get("sefe_gonderildi",0),
        "sahada":sum(by.get(x,0) for x in ["sahaya_atandi","kabul_edildi","yolda","yerinde","islemde"]),
        "onay":by.get("onay_bekliyor",0),
        "tamam":by.get("tamamlandi",0),
        "geri_bildirim_bekleyen":qs.filter(
            durum="tamamlandi",
            vatandas_bildirim_durumu__in=["bekliyor","tekrar_aranacak"],
        ).count(),
        "acil":qs.filter(oncelik="acil").exclude(durum__in=["tamamlandi","iptal"]).count(),
        "son_talepler":qs[:10],
        "map_points":map_points,
        "ilce_json":list(qs.values("ilce__ad").annotate(adet=Count("id")).order_by("-adet")[:8]),
        "tur_json":list(qs.values("is_turu__ad").annotate(adet=Count("id")).order_by("-adet")[:8]),
        "organizasyon_ozeti":organizasyon_ozeti,
        "sla_geciken":sla_geciken[:12],
        "sla_yaklasan":sla_yaklasan[:12],
        "sla_geciken_sayisi":len(sla_geciken),
        "sla_yaklasan_sayisi":len(sla_yaklasan),
        "sla_kurallari":SLA_ATAMA_DAKIKA,
        "abonelik_ozet":abonelik_ozet,
    })


@panel_required("admin")
def sistem_personel_saha(request):
    acik_saha_durumlari=["sahaya_atandi","kabul_edildi","yolda","yerinde","islemde","onay_bekliyor"]
    operasyon_durumlari=["kabul_edildi","yolda","yerinde","islemde"]

    tum_profiller=PersonelProfili.objects.select_related(
        "kullanici","rol"
    ).prefetch_related(
        "yetkili_ilceler","uzmanlik_is_turleri"
    ).order_by("rol__panel_tipi","rol__ad","kullanici__username")

    from abonelik.models import AmbarYetkisi
    ambar_yetki_map = {
        x.ilce_id: x for x in AmbarYetkisi.objects.select_related(
            "ilce", "ambar", "personel__kullanici", "personel__rol"
        ).filter(aktif=True)
    }

    ilce_kartlari=[]
    for ilce in Ilce.objects.filter(aktif=True).order_by("ad"):
        sefler=list(
            tum_profiller.filter(
                rol__panel_tipi="sef",
                yetkili_ilceler=ilce,
            ).distinct()
        )
        saha=list(
            tum_profiller.filter(
                rol__panel_tipi="saha",
                yetkili_ilceler=ilce,
            ).distinct()
        )

        aktif_saha_ids=set(
            Talep.objects.filter(
                ilce=ilce,
                durum__in=acik_saha_durumlari,
                sorumlu_saha__isnull=False,
            ).values_list("sorumlu_saha_id",flat=True)
        )

        for p in saha:
            p.aktif_gorev_sayisi=Talep.objects.filter(
                sorumlu_saha=p,
                durum__in=acik_saha_durumlari,
            ).count()
            p.aktif_operasyon_sayisi=Talep.objects.filter(
                sorumlu_saha=p,
                durum__in=operasyon_durumlari,
            ).count()
            p.operasyon_musait=(
                p.aktif and p.kullanici.is_active and p.musait
                and p.aktif_operasyon_sayisi==0
            )

        mevcut_ambar_yetkisi = ambar_yetki_map.get(ilce.id)
        for p in sefler:
            p.acik_talep_sayisi=Talep.objects.filter(
                sorumlu_koordinator=p
            ).exclude(
                durum__in=["tamamlandi","iptal"]
            ).count()
            p.ambar_sorumlusu = bool(mevcut_ambar_yetkisi and mevcut_ambar_yetkisi.personel_id == p.id)

        ilce_kartlari.append({
            "ilce":ilce,
            "sefler":sefler,
            "saha":saha,
            "aktif_operasyon":Talep.objects.filter(
                ilce=ilce,durum__in=operasyon_durumlari
            ).count(),
            "bekleyen":Talep.objects.filter(
                ilce=ilce,durum__in=["yeni","sefe_gonderildi"]
            ).count(),
            "musait_saha":sum(1 for p in saha if p.operasyon_musait),
            "ambar_yetkisi":mevcut_ambar_yetkisi,
        })

    superusers=User.objects.filter(
        is_superuser=True
    ).order_by("username")

    cagrimerkezi=list(
        tum_profiller.filter(
            rol__panel_tipi="185"
        )
    )

    bolgesiz=list(
        tum_profiller.exclude(
            rol__panel_tipi__in=["185","sef","saha"]
        )
    )

    return render(request,"system/personel_saha.html",{
        "ilce_kartlari":ilce_kartlari,
        "superusers":superusers,
        "cagrimerkezi":cagrimerkezi,
        "bolgesiz":bolgesiz,
        "toplam_kullanici":User.objects.count(),
        "aktif_kullanici":User.objects.filter(is_active=True).count(),
        "sef_sayisi":tum_profiller.filter(rol__panel_tipi="sef",aktif=True).count(),
        "saha_sayisi":tum_profiller.filter(rol__panel_tipi="saha",aktif=True).count(),
        "ambar_yetki_sayisi":len(ambar_yetki_map),
    })


@panel_required("admin")
@require_POST
def ambar_yetkisi_ata(request, ilce_id):
    """V54: İlçedeki şef/koordinatörlerden yalnız birini ambar sorumlusu yapar."""
    from abonelik.models import Ambar, AmbarYetkisi

    ilce = get_object_or_404(Ilce, pk=ilce_id, aktif=True)
    personel_id = request.POST.get("personel_id")
    personel = get_object_or_404(
        PersonelProfili.objects.select_related("kullanici", "rol").filter(
            pk=personel_id, aktif=True, rol__panel_tipi="sef", yetkili_ilceler=ilce
        ).distinct()
    )
    yerel, _ = Ambar.objects.get_or_create(
        kod=f"{slugify(ilce.ad)}-sayac-ambari",
        defaults={
            "ad": f"{ilce.ad} Sayaç Ambarı", "tur": "ilce", "ilce": ilce, "aktif": True,
            "aciklama": f"{ilce.ad} ilçesi yerel sayaç teslim, kontrol ve stok noktası.",
        },
    )
    if yerel.ilce_id != ilce.id or yerel.tur != "ilce":
        raise PermissionDenied("İlçe ambar kaydı bu yetki için uygun değil.")
    AmbarYetkisi.objects.update_or_create(
        ilce=ilce,
        defaults={"personel": personel, "ambar": yerel, "aktif": True},
    )
    messages.success(
        request, f"{ilce.ad} ambar sorumluluğu {personel.kullanici.get_full_name() or personel.kullanici.username} kullanıcısına verildi."
    )
    return redirect(f"{reverse('dashboard:sistem_personel_saha')}#ilce-{ilce.kod}")


@panel_required("admin")
def sistem_talepler(request):
    qs=Talep.objects.select_related("ilce","mahalle","is_turu","is_alt_turu","sorumlu_saha__kullanici")
    filtre=TalepFiltreForm(request.GET)
    if filtre.is_valid():
        c=filtre.cleaned_data
        if c["durum"]:qs=qs.filter(durum=c["durum"])
        if c["oncelik"]:qs=qs.filter(oncelik=c["oncelik"])
        if c["ilce"]:qs=qs.filter(ilce=c["ilce"])
        if c["is_turu"]:qs=qs.filter(is_turu=c["is_turu"])
        if c["ara"]:qs=qs.filter(Q(talep_no__icontains=c["ara"])|Q(telefon__icontains=c["ara"])|Q(vatandas_ad__icontains=c["ara"])|Q(vatandas_soyad__icontains=c["ara"]))
    talepler=_sla_ekle(list(qs[:500]))
    return render(request,"system/talepler.html",{"talepler":talepler,"filtre":filtre})

@panel_required("admin")
def sistem_kullanicilar(request):
    form=PersonelOlusturForm(request.POST or None)
    generated=request.session.pop("generated_accounts",None)

    if request.method=="POST" and request.POST.get("action")=="create" and form.is_valid():
        c=form.cleaned_data

        # Tek grid içindeki özel seçenek: tüm eksik organizasyonu algoritmayla tamamla.
        if c["kayit_alani"]=="auto":
            generated=create_missing_organization()
            for x in generated:
                user=User.objects.get(username=x["username"])
                log_yaz(
                    request,
                    "AUTO_ORGANIZATION_CREATE",
                    f"{x.get('ilce','')} / {x['rol']} hesabı otomatik oluşturuldu.",
                    varlik_turu="User",
                    varlik_id=user.pk,
                )
            request.session["generated_accounts"]=generated
            if generated:
                messages.success(
                    request,
                    f"Organizasyon tamamlandı: {len(generated)} eksik hesap oluşturuldu."
                )
            else:
                messages.success(
                    request,
                    "Organizasyon zaten tam; eksik koordinatör veya saha hesabı bulunmadı."
                )
            return redirect("dashboard:sistem_kullanicilar")

        role=c["rol_obj"]
        pwd=c["gecici_sifre"] or ("Isu#"+secrets.token_urlsafe(7))
        username=(c["kullanici_adi"] or "").strip()

        if User.objects.filter(username=username).exists():
            messages.error(request,"Bu kullanıcı adı zaten var.")
        else:
            with transaction.atomic():
                user=User.objects.create_user(
                    username=username,
                    password=pwd,
                    first_name=c["ad"],
                    last_name=c["soyad"],
                    email=c["eposta"],
                )
                p=PersonelProfili.objects.create(
                    kullanici=user,
                    rol=role,
                    telefon=c["telefon"],
                    sicil_no=c["sicil_no"] or None,
                    aktif=True,
                    musait=True,
                )
                p.yetkili_ilceler.set(c["yetkili_ilceler"])

                specs=list(c["uzmanlik_is_turleri"])
                if role.panel_tipi=="saha" and not specs:
                    bp=next(
                        (x for x in SAHA_BLUEPRINTS if x["kod"]==role.kod),
                        None
                    )
                    if bp and bp["is_turleri"]:
                        specs=list(
                            IsTuru.objects.filter(
                                kod__in=bp["is_turleri"],
                                aktif=True
                            )
                        )
                p.uzmanlik_is_turleri.set(specs)

            request.session["generated_accounts"]=[{
                "username":user.username,
                "password":pwd,
                "rol":p.rol.ad,
                "kanal":p.rol.get_calisma_kanali_display(),
                "ilce":", ".join(p.yetkili_ilceler.values_list("ad",flat=True)),
                "uzmanlik":", ".join(p.uzmanlik_is_turleri.values_list("ad",flat=True)) or "Genel",
            }]
            log_yaz(
                request,
                "USER_CREATE",
                f"{user.username} / {p.rol.ad} personeli oluşturuldu.",
                varlik_turu="User",
                varlik_id=user.pk,
            )
            messages.success(request,f"{p.rol.ad} personel kaydı oluşturuldu.")
            return redirect("dashboard:sistem_kullanicilar")

    users=PersonelProfili.objects.select_related(
        "kullanici","rol"
    ).prefetch_related(
        "yetkili_ilceler","uzmanlik_is_turleri"
    )
    return render(request,"system/kullanicilar.html",{
        "form":form,
        "personeller":users,
        "generated":generated,
        "saha_blueprints":SAHA_BLUEPRINTS,
    })

def _pwd():
    alphabet=string.ascii_letters+string.digits
    return "Isu#"+"".join(secrets.choice(alphabet) for _ in range(9))

@panel_required("admin")
def otomatik_hesaplar(request,tur):
    if request.method!="POST":
        raise PermissionDenied

    generated=[]

    if tur=="koordinator":
        rol=get_object_or_404(Rol,kod="koordinator")
        prefix="koord"
        for ilce in Ilce.objects.filter(aktif=True).order_by("ad"):
            if PersonelProfili.objects.filter(
                rol=rol,yetkili_ilceler=ilce,aktif=True
            ).exists():
                continue

            base=f"{prefix}_{slugify(ilce.ad).replace('-','_')}"
            username=base
            n=2
            while User.objects.filter(username=username).exists():
                username=f"{base}_{n}"
                n+=1

            pwd=_pwd()
            user=User.objects.create_user(
                username=username,password=pwd,
                first_name=ilce.ad,last_name=rol.ad
            )
            p=PersonelProfili.objects.create(
                kullanici=user,rol=rol,aktif=True,musait=True
            )
            p.yetkili_ilceler.add(ilce)
            generated.append({
                "username":username,"password":pwd,
                "rol":rol.ad,"ilce":ilce.ad,"uzmanlik":"İlçe koordinasyonu"
            })
            log_yaz(
                request,"AUTO_USER_CREATE",
                f"{ilce.ad} için koordinatör hesabı otomatik oluşturuldu.",
                varlik_turu="User",varlik_id=user.pk
            )

    elif tur in ["saha","saha-tum"]:
        generated=create_missing_field_teams()
        for x in generated:
            user=User.objects.get(username=x["username"])
            log_yaz(
                request,"AUTO_FIELD_TEAM_CREATE",
                f"{x['ilce']} için {x['rol']} hesabı otomatik oluşturuldu.",
                varlik_turu="User",varlik_id=user.pk
            )
    else:
        raise PermissionDenied

    request.session["generated_accounts"]=generated
    if generated:
        messages.success(
            request,
            f"{len(generated)} eksik hesap otomatik oluşturuldu."
        )
    else:
        messages.success(
            request,
            "Eksik hesap yok; mevcut organizasyon kuralları karşılıyor."
        )
    return redirect("dashboard:sistem_kullanicilar")

@panel_required("admin")
@require_POST
def personel_toggle(request,pk):
    p=get_object_or_404(PersonelProfili,pk=pk)

    if p.aktif:
        saha_acik=p.saha_talepleri.exclude(
            durum__in=["tamamlandi","iptal"]
        ).exists()
        koord_acik=p.koordinator_talepleri.exclude(
            durum__in=["tamamlandi","iptal"]
        ).exists()

        if saha_acik or koord_acik:
            messages.error(
                request,
                "Aktif talep sorumluluğu bulunan personel pasife alınamaz. Önce talepleri tamamlayın veya yeniden atayın."
            )
            return redirect("dashboard:sistem_kullanicilar")

    p.aktif=not p.aktif
    p.kullanici.is_active=p.aktif
    p.save(update_fields=["aktif"])
    p.kullanici.save(update_fields=["is_active"])
    log_yaz(
        request,
        "USER_TOGGLE",
        f"{p.kullanici.username}: {'aktif' if p.aktif else 'pasif'}",
        varlik_turu="Personel",
        varlik_id=p.pk
    )
    messages.success(
        request,
        f"{p.kullanici.username} {'aktifleştirildi' if p.aktif else 'pasife alındı'}."
    )
    return redirect("dashboard:sistem_kullanicilar")

@panel_required("admin")
def sistem_roller(request):
    form=RolForm(request.POST or None)
    if request.method=="POST" and form.is_valid():
        r=form.save();log_yaz(request,"ROLE_SAVE",f"{r.ad} rolü kaydedildi.",varlik_turu="Rol",varlik_id=r.pk)
        messages.success(request,"Rol kaydedildi.");return redirect("dashboard:sistem_roller")
    return render(request,"system/roller.html",{"form":form,"roller":Rol.objects.select_related("parent")})

@panel_required("admin")
def sistem_kurallar(request):
    form=RolAtamaKuraliForm(request.POST or None)
    if request.method=="POST" and form.is_valid():
        k=form.save();log_yaz(request,"RULE_SAVE",f"{k} atama kuralı kaydedildi.",varlik_turu="RolAtamaKurali",varlik_id=k.pk)
        messages.success(request,"Atama kuralı kaydedildi.");return redirect("dashboard:sistem_kurallar")
    return render(request,"system/kurallar.html",{"form":form,"kurallar":RolAtamaKurali.objects.select_related("kaynak_rol","hedef_rol")})

@panel_required("admin")
def sistem_is_turleri(request):
    form=IsTuruForm(request.POST or None,prefix="tur");alt=IsAltTuruForm(request.POST or None,prefix="alt")
    action=request.POST.get("action") if request.method=="POST" else ""
    if request.method=="POST" and action=="tur" and form.is_valid():
        x=form.save();log_yaz(request,"WORKTYPE_SAVE",f"{x.ad} iş türü kaydedildi.",varlik_turu="IsTuru",varlik_id=x.pk);return redirect("dashboard:sistem_is_turleri")
    if request.method=="POST" and action=="alt" and alt.is_valid():
        x=alt.save();log_yaz(request,"SUBTYPE_SAVE",f"{x} iş alt türü kaydedildi.",varlik_turu="IsAltTuru",varlik_id=x.pk);return redirect("dashboard:sistem_is_turleri")
    if request.method=="POST" and action=="foto_kurali":
        alt_id=request.POST.get("alt_id")
        x=get_object_or_404(IsAltTuru,pk=alt_id)
        try:
            adet=int(request.POST.get("zorunlu_fotograf_sayisi") or 1)
        except ValueError:
            adet=0
        if not 1<=adet<=8:
            messages.error(request,"Zorunlu fotoğraf sayısı 1 ile 8 arasında olmalıdır.")
            return redirect("dashboard:sistem_is_turleri")
        etiketler="\n".join(
            z.strip() for z in (request.POST.get("fotograf_etiketleri") or "").splitlines() if z.strip()
        )
        x.zorunlu_fotograf_sayisi=adet
        x.fotograf_etiketleri=etiketler
        x.save(update_fields=["zorunlu_fotograf_sayisi","fotograf_etiketleri"])
        log_yaz(request,"SUBTYPE_PHOTO_RULE",f"{x}: {adet} zorunlu fotoğraf.",varlik_turu="IsAltTuru",varlik_id=x.pk)
        messages.success(request,f"{x.ad} için fotoğraf kuralı güncellendi.")
        return redirect("dashboard:sistem_is_turleri")
    return render(request,"system/is_turleri.html",{"form":form,"alt_form":alt,"turler":IsTuru.objects.prefetch_related("alt_turler")})

@panel_required("admin")
def sistem_adresler(request):
    return render(request,"system/adresler.html",{
        "csv_form":AdresCsvForm(),"ilceler":Ilce.objects.annotate(mahalle_sayisi=Count("mahalleler")),
        "mahalle_sayisi":Mahalle.objects.count(),"yol_sayisi":Yol.objects.count()
    })

@panel_required("admin")
def adres_csv_import(request):
    if request.method!="POST":raise PermissionDenied
    form=AdresCsvForm(request.POST,request.FILES)
    if not form.is_valid():messages.error(request,"CSV dosyası seçilemedi.");return redirect("dashboard:sistem_adresler")
    raw=form.cleaned_data["dosya"].read()
    try:text=raw.decode("utf-8-sig")
    except UnicodeDecodeError:text=raw.decode("cp1254")
    reader=csv.DictReader(io.StringIO(text))
    if not {"ilce","mahalle","yol"}.issubset(set(reader.fieldnames or [])):
        messages.error(request,"CSV'de ilce, mahalle ve yol kolonları zorunlu.");return redirect("dashboard:sistem_adresler")
    mc=yc=0
    with transaction.atomic():
        for row in reader:
            ia=(row.get("ilce") or "").strip();ma=(row.get("mahalle") or "").strip();ya=(row.get("yol") or "").strip()
            if not ia or not ma or not ya:continue
            i,_=Ilce.objects.get_or_create(ad=ia);m,cm=Mahalle.objects.get_or_create(ilce=i,ad=ma);mc+=int(cm)
            defaults={"tur":(row.get("tur") or "sokak").strip().lower()}
            if row.get("lat"):defaults["merkez_lat"]=row["lat"].strip()
            if row.get("lng"):defaults["merkez_lng"]=row["lng"].strip()
            _,cy=Yol.objects.get_or_create(mahalle=m,ad=ya,defaults=defaults);yc+=int(cy)
    log_yaz(request,"ADDRESS_IMPORT",f"Adres CSV içe aktarıldı: {mc} mahalle, {yc} yol.",varlik_turu="Adres")
    messages.success(request,f"İçe aktarma tamamlandı: {mc} yeni mahalle, {yc} yeni yol.");return redirect("dashboard:sistem_adresler")

def _memnuniyet_notunu_ayristir(not_metni):
    """V17-V19 geri bildirim notlarını raporlanabilir alanlara ayırır."""
    data={
        "memnuniyet":"",
        "islem_suresi":"",
        "sorun_cozuldu":"",
        "hizmet_hizi":"",
        "bilgilendirme":"",
        "personel_iletisimi":"",
        "genel_puan":"",
        "gorusme_notu":"",
    }
    if not not_metni:
        return data

    for parca in [x.strip() for x in not_metni.split(" | ") if x.strip()]:
        if ":" not in parca:
            continue
        anahtar,deger=parca.split(":",1)
        anahtar=anahtar.strip().casefold()
        deger=deger.strip()

        if anahtar in {"genel memnuniyet","memnuniyet"}:
            data["memnuniyet"]=deger
        elif anahtar=="işlem süresi":
            data["islem_suresi"]=deger
        elif anahtar in {"anket / sorun çözüldü mü","sorun çözüldü mü"}:
            data["sorun_cozuldu"]=deger
        elif anahtar in {"anket / hizmet hızı","hizmet hızı"}:
            data["hizmet_hizi"]=deger
        elif anahtar in {"anket / bilgilendirme","bilgilendirme"}:
            data["bilgilendirme"]=deger
        elif anahtar in {"anket / personel iletişimi","personel iletişimi"}:
            data["personel_iletisimi"]=deger
        elif anahtar in {"anket / genel hizmet puanı","genel hizmet puanı"}:
            data["genel_puan"]=deger
        elif anahtar=="görüşme notu":
            data["gorusme_notu"]=deger
    return data


@panel_required("admin")
def sistem_memnuniyet(request):
    qs=VatandasAramaKaydi.objects.filter(
        sonuc="bilgilendirildi"
    ).select_related(
        "talep","talep__ilce","talep__is_turu","kullanici"
    ).order_by("-tarih")

    ara=(request.GET.get("ara") or "").strip()
    ilce_id=(request.GET.get("ilce") or "").strip()

    if ara:
        qs=qs.filter(
            Q(talep__talep_no__icontains=ara)
            |Q(talep__vatandas_ad__icontains=ara)
            |Q(talep__vatandas_soyad__icontains=ara)
            |Q(kullanici__username__icontains=ara)
        )
    if ilce_id.isdigit():
        qs=qs.filter(talep__ilce_id=int(ilce_id))

    satirlar=[]
    puanlar=[]
    memnuniyet_sayisi=0
    iyi_sayisi=0
    anket_sayisi=0
    dikkat_sayisi=0

    for kayit in qs[:1000]:
        d=_memnuniyet_notunu_ayristir(kayit.not_metni)
        anket_var=any([
            d["sorun_cozuldu"],d["hizmet_hizi"],d["bilgilendirme"],
            d["personel_iletisimi"],d["genel_puan"],
        ])
        if anket_var:
            anket_sayisi+=1

        if d["memnuniyet"]:
            memnuniyet_sayisi+=1
            if d["memnuniyet"].casefold()=="iyi":
                iyi_sayisi+=1

        puan_num=None
        if d["genel_puan"]:
            try:
                puan_num=int(d["genel_puan"].split("/",1)[0].strip())
            except (TypeError,ValueError):
                puan_num=None
            if puan_num is not None and 1<=puan_num<=5:
                puanlar.append(puan_num)

        dikkat=(
            d["memnuniyet"].casefold()=="kötü"
            or d["sorun_cozuldu"].casefold()=="hayır"
            or d["hizmet_hizi"].casefold()=="yavaş"
            or d["bilgilendirme"].casefold()=="yetersiz"
            or d["personel_iletisimi"].casefold()=="kötü"
            or (puan_num is not None and puan_num<=2)
        )
        if dikkat:
            dikkat_sayisi+=1

        satirlar.append({
            "kayit":kayit,
            "memnuniyet":d["memnuniyet"],
            "islem_suresi":d["islem_suresi"],
            "sorun_cozuldu":d["sorun_cozuldu"],
            "hizmet_hizi":d["hizmet_hizi"],
            "bilgilendirme":d["bilgilendirme"],
            "personel_iletisimi":d["personel_iletisimi"],
            "genel_puan":d["genel_puan"],
            "gorusme_notu":d["gorusme_notu"],
            "anket_var":anket_var,
            "dikkat":dikkat,
        })

    toplam=len(satirlar)
    ortalama_puan=round(sum(puanlar)/len(puanlar),1) if puanlar else None
    memnuniyet_orani=round((iyi_sayisi/memnuniyet_sayisi)*100) if memnuniyet_sayisi else 0
    anket_katilim=round((anket_sayisi/toplam)*100) if toplam else 0

    return render(request,"system/memnuniyet.html",{
        "satirlar":satirlar,
        "toplam":toplam,
        "anket_sayisi":anket_sayisi,
        "anket_katilim":anket_katilim,
        "ortalama_puan":ortalama_puan,
        "memnuniyet_orani":memnuniyet_orani,
        "dikkat_sayisi":dikkat_sayisi,
        "ilceler":Ilce.objects.order_by("ad"),
        "secili_ilce":ilce_id,
        "ara":ara,
    })


@panel_required("admin")
def sistem_geri_bildirimler(request):
    return render(request,"system/geri_bildirimler.html",{"items":GeriBildirim.objects.select_related("talep","kullanici")[:500]})

@panel_required("admin")
def sistem_loglar(request):
    qs=IslemLogu.objects.select_related("kullanici","talep");ara=(request.GET.get("ara") or "").strip()
    if ara:qs=qs.filter(Q(aciklama__icontains=ara)|Q(islem__icontains=ara)|Q(talep__talep_no__icontains=ara)|Q(kullanici__username__icontains=ara))
    return render(request,"system/loglar.html",{"loglar":qs[:1000],"ara":ara})



def _otomatik_rapor_context():
    """
    V16 raporları kullanıcıdan manuel veri istemez.
    Talep ve işlem kayıtları ne durumdaysa rapor doğrudan veritabanından üretilir.
    """
    qs=Talep.objects.select_related(
        "ilce","mahalle","yol","is_turu","is_alt_turu",
        "sorumlu_koordinator__kullanici","sorumlu_saha__kullanici",
    ).order_by("-olusturulma_tarihi")

    toplam=qs.count()
    yeni=qs.filter(durum__in=["yeni","sefe_gonderildi"]).count()
    sahada=qs.filter(
        durum__in=["sahaya_atandi","kabul_edildi","yolda","yerinde","islemde"]
    ).count()
    onay=qs.filter(durum="onay_bekliyor").count()
    tamam=qs.filter(durum="tamamlandi").count()
    acil=qs.filter(oncelik="acil").exclude(
        durum__in=["tamamlandi","iptal"]
    ).count()
    geri_bekleyen=qs.filter(
        durum="tamamlandi",
        vatandas_bildirim_durumu__in=["bekliyor","tekrar_aranacak"],
    ).count()
    geri_tamam=qs.filter(
        vatandas_bildirim_durumu="bilgilendirildi"
    ).count()

    # İlçe -> mahalle -> cadde/sokak ölçeklerinde otomatik yoğunluk.
    ilce_rows=list(
        qs.values("ilce__ad")
        .annotate(
            adet=Count("id"),
            aktif=Count("id",filter=~Q(durum__in=["tamamlandi","iptal"])),
            tamam=Count("id",filter=Q(durum="tamamlandi")),
            acil=Count("id",filter=Q(oncelik="acil") & ~Q(durum__in=["tamamlandi","iptal"])),
            son_talep=Max("guncellenme_tarihi"),
        )
        .order_by("-adet","ilce__ad")
    )

    mahalle_rows=list(
        qs.values("mahalle__ad","ilce__ad")
        .annotate(
            adet=Count("id"),
            aktif=Count("id",filter=~Q(durum__in=["tamamlandi","iptal"])),
            tamam=Count("id",filter=Q(durum="tamamlandi")),
            acil=Count("id",filter=Q(oncelik="acil") & ~Q(durum__in=["tamamlandi","iptal"])),
            son_talep=Max("guncellenme_tarihi"),
        )
        .order_by("-adet","ilce__ad","mahalle__ad")
    )

    yol_rows=list(
        qs.values("yol__ad","yol__tur","mahalle__ad","ilce__ad")
        .annotate(
            adet=Count("id"),
            aktif=Count("id",filter=~Q(durum__in=["tamamlandi","iptal"])),
            tamam=Count("id",filter=Q(durum="tamamlandi")),
            acil=Count("id",filter=Q(oncelik="acil") & ~Q(durum__in=["tamamlandi","iptal"])),
            son_talep=Max("guncellenme_tarihi"),
        )
        .order_by("-adet","ilce__ad","mahalle__ad","yol__ad")
    )

    # Görsel yoğunluk çubuğu için en yüksek sayıya göre ölçekle.
    def _bar_orani(rows):
        max_adet=max([x["adet"] for x in rows],default=0)
        for x in rows:
            x["oran"]=round((x["adet"]/max_adet)*100,1) if max_adet else 0
        return rows

    _bar_orani(ilce_rows)
    _bar_orani(mahalle_rows)
    _bar_orani(yol_rows)

    # Eski Sistem Yönetim panelindeki harita yapısı:
    # her talep ayrı pin; pin rengi durum, iç çekirdek aciliyet bilgisidir.
    map_points=[]
    for t in qs.exclude(lat__isnull=True).exclude(lng__isnull=True)[:500]:
        map_points.append({
            "id":t.id,
            "no":t.talep_no,
            "lat":float(t.lat),
            "lng":float(t.lng),
            "durum":t.durum,
            "durum_label":t.get_durum_display(),
            "oncelik":t.oncelik,
            "oncelik_label":t.get_oncelik_display(),
            "ilce":t.ilce.ad,
            "mahalle":t.mahalle.ad,
            "yol":t.yol.ad,
            "tur":t.is_turu.ad,
            "alt_tur":t.is_alt_turu.ad,
            "guncelleme":timezone.localtime(t.guncellenme_tarihi).strftime("%d.%m.%Y %H:%M"),
        })

    tamamlanma_sureleri=[]
    for bas,bit in qs.filter(
        durum="tamamlandi",
        tamamlanma_tarihi__isnull=False,
    ).values_list("olusturulma_tarihi","tamamlanma_tarihi"):
        if bas and bit and bit>=bas:
            tamamlanma_sureleri.append((bit-bas).total_seconds()/3600)
    ort_sure=round(
        sum(tamamlanma_sureleri)/len(tamamlanma_sureleri),1
    ) if tamamlanma_sureleri else 0

    latest_talep=qs.aggregate(son=Max("guncellenme_tarihi"))["son"]
    latest_log=IslemLogu.objects.aggregate(son=Max("tarih"))["son"]
    version="|".join([
        str(toplam),
        latest_talep.isoformat() if latest_talep else "0",
        latest_log.isoformat() if latest_log else "0",
    ])

    return {
        "toplam":toplam,
        "yeni":yeni,
        "sahada":sahada,
        "onay":onay,
        "tamam":tamam,
        "acil":acil,
        "geri_bekleyen":geri_bekleyen,
        "geri_tamam":geri_tamam,
        "ortalama_tamamlanma_saati":ort_sure,
        "tamamlanma_orani":round((tamam/toplam)*100,1) if toplam else 0,
        "geri_donus_orani":round((geri_tamam/tamam)*100,1) if tamam else 0,
        "ilce_rows":ilce_rows,
        "mahalle_rows":mahalle_rows,
        "yol_rows":yol_rows,
        "en_yogun_ilce":ilce_rows[0] if ilce_rows else None,
        "en_yogun_mahalle":mahalle_rows[0] if mahalle_rows else None,
        "en_yogun_yol":yol_rows[0] if yol_rows else None,
        "map_points":map_points,
        "is_listesi":list(qs[:60]),
        "rapor_version":version,
        "rapor_zamani":timezone.localtime(),
    }



# ---------------------------------------------------------------------
# V23 — Detaylı Rapor Merkezi
# Sistem yöneticisi tarih aralığını seçer. Varsayılan dönem içinde
# bulunulan aydır. Aynı filtre ekrandaki yatay tabloya ve dışa aktarıma
# uygulanır.
# ---------------------------------------------------------------------

DETAYLI_RAPOR_KOLONLARI = [
    ("talep_no", "Talep No"),
    ("is_emri_no", "İş Emri No"),
    ("talep_baslangic", "Talep Tarihi / Saat"),
    ("talep_adi", "Talep Adı"),
    ("vatandas", "Vatandaş"),
    ("telefon", "Telefon"),
    ("eposta", "E-posta"),
    ("ilce", "İlçe"),
    ("mahalle", "Mahalle"),
    ("yol", "Cadde / Sokak"),
    ("kapi_no", "Kapı No"),
    ("adres", "Tam Adres / Tarif"),
    ("koordinat", "Talep Koordinatı"),
    ("is_turu", "İş Türü"),
    ("is_alt_turu", "İş Alt Türü"),
    ("oncelik", "Öncelik"),
    ("durum", "Güncel Durum"),
    ("talep_eden_birim", "Talep Eden Birim"),
    ("gonderen_birim", "İş Emri Gönderen Birim"),
    ("koordinator", "Alan Şef / Koordinatör"),
    ("koordinator_rol", "Şef / Koordinatör Rolü"),
    ("koordinator_tarihi", "Şefe / Koordinatöre Düştü"),
    ("saha", "Gönderilen Saha"),
    ("saha_rol", "Saha Rolü / Uzmanlığı"),
    ("saha_atama_tarihi", "Sahaya Atama Tarihi"),
    ("sla_atama_suresi", "Sahaya Atama Süresi"),
    ("saha_kabul_tarihi", "Saha Kabul"),
    ("yola_cikis_tarihi", "Yola Çıkış"),
    ("adrese_ulasma_tarihi", "Adrese Ulaşma"),
    ("mudahale_baslama_tarihi", "Müdahale Başlama"),
    ("saha_bitirdi", "Saha İşi Bitirdi"),
    ("sef_onaylayan", "Şef Onaylayan"),
    ("sef_onay_tarihi", "Şef Onay Tarihi"),
    ("tekrar_baslatildi", "Tekrar Başlatıldı mı?"),
    ("tekrar_sayisi", "İade / Tekrar Sayısı"),
    ("tamamlanma_tarihi", "Teknik Tamamlanma"),
    ("is_emri_suresi", "İş Emri Süresi"),
    ("toplam_cozum_suresi", "Toplam Çözüm Süresi"),
    ("sla_atama_hedefi", "SLA Hedefi"),
    ("sla_atama_sonucu", "SLA Sonucu"),
    ("gps_dogrulama", "GPS Doğrulama"),
    ("gps_mesafe", "GPS Mesafe"),
    ("zorunlu_gorsel", "Zorunlu Görsel Adedi"),
    ("gorsel_durumu", "İş Emri Görselleri"),
    ("geri_donus_durumu", "Geri Dönüş"),
    ("geri_donus_tarihi", "Geri Dönüş Tarihi"),
    ("geri_donus_personeli", "Geri Dönüş Personeli"),
    ("gorusme_sonucu", "Son Görüşme Sonucu"),
    ("genel_memnuniyet", "Memnuniyet"),
    ("anket_puan", "Anket Puanı"),
    ("talep_aciklama", "Talep Açıklaması"),
    ("saha_sonuc", "Saha Sonuç Notu"),
    ("gorusme_notu", "Görüşme Notu"),
]


def _rapor_kullanici_adi(user):
    if not user:
        return "-"
    tam=(user.get_full_name() or "").strip()
    return tam or user.username


def _rapor_dt(dt):
    if not dt:
        return "-"
    return timezone.localtime(dt).strftime("%d.%m.%Y %H:%M")


def _rapor_tarih_araligi(request):
    bugun=timezone.localdate()
    varsayilan_baslangic=bugun.replace(day=1)

    raw_baslangic=(request.GET.get("baslangic") or "").strip()
    raw_bitis=(request.GET.get("bitis") or "").strip()

    try:
        baslangic=date.fromisoformat(raw_baslangic) if raw_baslangic else varsayilan_baslangic
    except ValueError:
        baslangic=varsayilan_baslangic

    try:
        bitis=date.fromisoformat(raw_bitis) if raw_bitis else bugun
    except ValueError:
        bitis=bugun

    if bitis < baslangic:
        baslangic,bitis=bitis,baslangic

    return baslangic,bitis


def _detayli_rapor_queryset(baslangic,bitis):
    return Talep.objects.filter(
        olusturulma_tarihi__date__gte=baslangic,
        olusturulma_tarihi__date__lte=bitis,
    ).select_related(
        "ilce","mahalle","yol","is_turu","is_alt_turu",
        "olusturan","olusturan__personel_profili__rol",
        "sorumlu_koordinator__kullanici","sorumlu_koordinator__rol",
        "sorumlu_saha__kullanici","sorumlu_saha__rol",
        "sef_onaylayan","vatandas_bildirim_yapan","abone","is_emri",
    ).prefetch_related(
        "geri_bildirimler","loglar","vatandas_arama_kayitlari","is_emri__fotograflar"
    ).order_by("-olusturulma_tarihi")


def _detayli_rapor_satiri(t):
    hareketler=sorted(list(t.geri_bildirimler.all()),key=lambda x:x.tarih)
    loglar=list(t.loglar.all())
    aramalar=sorted(list(t.vatandas_arama_kayitlari.all()),key=lambda x:x.tarih,reverse=True)

    def ilk_hareket(durum):
        return next((x for x in hareketler if x.durum==durum),None)

    koord_hareket=ilk_hareket("sefe_gonderildi")
    saha_hareket=ilk_hareket("sahaya_atandi")
    kabul_hareket=ilk_hareket("kabul_edildi")
    yolda_hareket=ilk_hareket("yolda")
    yerinde_hareket=ilk_hareket("yerinde")
    mudahale_hareket=ilk_hareket("islemde")
    iade_sayisi=sum(1 for x in loglar if x.islem=="CHIEF_RETURN")

    sla_hedef=SLA_ATAMA_DAKIKA.get(t.oncelik,120)
    sla_baslangic=koord_hareket.tarih if koord_hareket else t.olusturulma_tarihi
    if saha_hareket:
        sla_gercek=max(0,int((saha_hareket.tarih-sla_baslangic).total_seconds()//60))
        if sla_gercek <= sla_hedef:
            sla_sonuc="Zamanında"
        else:
            sla_sonuc=f"Aşıldı (+{sla_gercek-sla_hedef} dk)"
        sla_sure=f"{sla_gercek} dk"
    elif t.sorumlu_saha_id:
        # Eski kayıtlarda hareket kaydı bulunmayabilir.
        sla_sure="-"
        sla_sonuc="Atandı / tarih kaydı yok"
    elif t.durum in ("yeni","sefe_gonderildi"):
        sla_canli=_sla_bilgisi(t)
        sla_sure=f"{sla_canli['gecen_dakika']} dk (devam ediyor)"
        if sla_canli["durum"]=="asildi":
            sla_sonuc=f"SLA AŞILDI (+{sla_canli['asma_dakika']} dk)"
        elif sla_canli["durum"]=="yaklasiyor":
            sla_sonuc=f"Yaklaşıyor ({sla_canli['kalan_dakika']} dk kaldı)"
        else:
            sla_sonuc=f"Süre içinde ({sla_canli['kalan_dakika']} dk kaldı)"
    else:
        sla_sure="-"
        sla_sonuc="-"

    son_arama=aramalar[0] if aramalar else None
    son_tamamlanan=next((x for x in aramalar if x.sonuc=="bilgilendirildi"),None)
    anket=_memnuniyet_notunu_ayristir(
        son_tamamlanan.not_metni if son_tamamlanan else ""
    )

    adres_parcalari=[
        t.ilce.ad if t.ilce_id else "",
        t.mahalle.ad if t.mahalle_id else "",
        t.yol.ad if t.yol_id else "",
        f"No: {t.kapi_no}" if t.kapi_no else "",
        t.adres_aciklama or "",
    ]
    tam_adres=" / ".join([x for x in adres_parcalari if x]) or "-"

    koordinator=t.sorumlu_koordinator
    saha=t.sorumlu_saha

    try:
        emri=t.is_emri
    except IsEmri.DoesNotExist:
        emri=None

    is_emri_dk=None
    if emri and emri.atama_tarihi and emri.sef_onay_tarihi:
        is_emri_dk=(emri.sef_onay_tarihi-emri.atama_tarihi).total_seconds()/60

    cozum_dk=None
    if t.tamamlanma_tarihi and t.olusturulma_tarihi:
        cozum_dk=(t.tamamlanma_tarihi-t.olusturulma_tarihi).total_seconds()/60

    veri={
        "talep_no":t.talep_no,
        "is_emri_no":emri.is_emri_no if emri else "-",
        "gonderen_birim":emri.gonderen_birim if emri else "-",
        "abone_no":t.abone.abone_no if t.abone_id else "-",
        "talep_baslangic":_rapor_dt(t.olusturulma_tarihi),
        "talep_adi":(
            f"{t.is_turu.ad} / {t.is_alt_turu.ad}"
            if t.is_turu_id and t.is_alt_turu_id
            else (t.is_turu.ad if t.is_turu_id else "-")
        ),
        "vatandas":f"{t.vatandas_ad} {t.vatandas_soyad}".strip(),
        "telefon":t.telefon or "-",
        "eposta":t.eposta or "-",
        "ilce":t.ilce.ad if t.ilce_id else "-",
        "mahalle":t.mahalle.ad if t.mahalle_id else "-",
        "yol":t.yol.ad if t.yol_id else "-",
        "kapi_no":t.kapi_no or "-",
        "adres":tam_adres,
        "koordinat":f"{t.lat}, {t.lng}" if t.lat is not None and t.lng is not None else "-",
        "is_turu":t.is_turu.ad if t.is_turu_id else "-",
        "is_alt_turu":t.is_alt_turu.ad if t.is_alt_turu_id else "-",
        "talep_aciklama":t.aciklama or "-",
        "oncelik":t.get_oncelik_display(),
        "durum":t.get_durum_display(),
        "talep_eden_birim":_rapor_acan_birim(t),
        "olusturan":_rapor_kullanici_adi(t.olusturan),
        "koordinator":_rapor_kullanici_adi(koordinator.kullanici) if koordinator else "-",
        "koordinator_rol":koordinator.rol.ad if koordinator and koordinator.rol_id else "-",
        "koordinator_tarihi":_rapor_dt(koord_hareket.tarih if koord_hareket else None),
        "saha":_rapor_kullanici_adi(saha.kullanici) if saha else "-",
        "saha_rol":saha.rol.ad if saha and saha.rol_id else "-",
        "saha_atama_tarihi":_rapor_dt(saha_hareket.tarih if saha_hareket else None),
        "saha_kabul_tarihi":_rapor_dt(kabul_hareket.tarih if kabul_hareket else (emri.kabul_tarihi if emri else None)),
        "yola_cikis_tarihi":_rapor_dt(yolda_hareket.tarih if yolda_hareket else (emri.yola_cikis_tarihi if emri else None)),
        "adrese_ulasma_tarihi":_rapor_dt(yerinde_hareket.tarih if yerinde_hareket else (emri.adrese_ulasma_tarihi if emri else None)),
        "mudahale_baslama_tarihi":_rapor_dt(mudahale_hareket.tarih if mudahale_hareket else (emri.mudahale_baslama_tarihi if emri else None)),
        "gps_dogrulama":("Doğrulandı" if emri and emri.gps_dogrulandi else ("Kaydedildi" if emri and emri.gps_dogrulama_tarihi else "-")),
        "gps_mesafe":(f"{emri.gps_mesafe_m} m" if emri and emri.gps_mesafe_m is not None else "-"),
        "zorunlu_gorsel":str(t.is_alt_turu.zorunlu_fotograf_sayisi if t.is_alt_turu_id else "-"),
        "gorsel_durumu":(
            f"{emri.fotograflar.count()}/{t.is_alt_turu.zorunlu_fotograf_sayisi} tamam"
            if emri and t.is_alt_turu_id else "-"
        ),
        "sla_atama_hedefi":f"{sla_hedef} dk",
        "sla_atama_suresi":sla_sure,
        "sla_atama_sonucu":sla_sonuc,
        "saha_bitirdi":_rapor_dt(t.saha_tamam_bildirim_tarihi),
        "saha_sonuc":t.saha_sonuc_notu or "-",
        "sef_onaylayan":_rapor_kullanici_adi(t.sef_onaylayan),
        "sef_onay_tarihi":_rapor_dt(t.sef_onay_tarihi),
        "tekrar_baslatildi":"Evet" if iade_sayisi else "Hayır",
        "tekrar_sayisi":str(iade_sayisi),
        "tamamlanma_tarihi":_rapor_dt(t.tamamlanma_tarihi),
        "is_emri_suresi":_sure_etiketi(is_emri_dk),
        "toplam_cozum_suresi":_sure_etiketi(cozum_dk),
        "geri_donus_durumu":t.get_vatandas_bildirim_durumu_display(),
        "geri_donus_tarihi":_rapor_dt(t.vatandas_bildirim_tarihi),
        "geri_donus_personeli":_rapor_kullanici_adi(t.vatandas_bildirim_yapan),
        "gorusme_sonucu":son_arama.get_sonuc_display() if son_arama else "-",
        "genel_memnuniyet":anket.get("memnuniyet") or "-",
        "islem_suresi":anket.get("islem_suresi") or "-",
        "anket_cozum":anket.get("sorun_cozuldu") or "-",
        "anket_hiz":anket.get("hizmet_hizi") or "-",
        "anket_bilgilendirme":anket.get("bilgilendirme") or "-",
        "anket_puan":anket.get("genel_puan") or "-",
        "gorusme_notu":anket.get("gorusme_notu") or "-",
    }
    return veri



def _sure_etiketi(dakika):
    if dakika is None:
        return "-"
    dakika=max(0,int(round(dakika)))
    if dakika>=1440:
        gun=dakika//1440
        kalan=dakika%1440
        saat=kalan//60
        dk=kalan%60
        return f"{gun} gün {saat} sa {dk} dk"
    if dakika>=60:
        return f"{dakika//60} sa {dakika%60} dk"
    return f"{dakika} dk"


def _rapor_acan_birim(talep):
    """Genel icmalde talebi açan organizasyon birimini üretir."""
    user=talep.olusturan
    if not user:
        return "-"
    try:
        profil=user.personel_profili
    except PersonelProfili.DoesNotExist:
        return "Sistem Yönetimi" if user.is_superuser else "Diğer"

    rol=profil.rol
    if not rol:
        return "Diğer"
    if rol.panel_tipi=="185":
        return "185 Çağrı Merkezi"
    if rol.panel_tipi=="admin":
        return "Sistem Yönetimi"
    return rol.ad or "Diğer"


def _rapor_icmal_hazirla(baslangic,bitis):
    talepler=list(_detayli_rapor_queryset(baslangic,bitis))
    is_emirleri=list(
        IsEmri.objects.filter(
            talep__olusturulma_tarihi__date__gte=baslangic,
            talep__olusturulma_tarihi__date__lte=bitis,
        ).select_related(
            "talep__ilce","talep__is_turu",
            "atanan_saha__kullanici","atanan_saha__rol",
        ).order_by("-atama_tarihi")
    )

    acik_emir=[e for e in is_emirleri if e.durum not in ("tamamlandi","iptal")]
    tamam_emir=[e for e in is_emirleri if e.durum=="tamamlandi"]

    # İş emri atama -> şef onayı, operasyonun gerçek iş-emri süresi.
    is_emri_sureleri=[
        (e.sef_onay_tarihi-e.atama_tarihi).total_seconds()/60
        for e in tamam_emir
        if e.sef_onay_tarihi and e.atama_tarihi
    ]
    ort_is_emri_dk=(
        sum(is_emri_sureleri)/len(is_emri_sureleri)
        if is_emri_sureleri else None
    )

    # Talep oluşturma -> teknik tamamlanma, uçtan uca çözüm süresi.
    cozum_sureleri=[
        (t.tamamlanma_tarihi-t.olusturulma_tarihi).total_seconds()/60
        for t in talepler
        if t.durum=="tamamlandi" and t.tamamlanma_tarihi
    ]
    ort_cozum_dk=(
        sum(cozum_sureleri)/len(cozum_sureleri)
        if cozum_sureleri else None
    )

    # Şefe/koordinatöre düşme -> sahaya atama süresi ve SLA.
    atama_sureleri=[]
    sla_asan=0
    sla_yaklasan=0
    for t in talepler:
        hareketler=sorted(list(t.geri_bildirimler.all()),key=lambda x:x.tarih)
        koord=next((x for x in hareketler if x.durum=="sefe_gonderildi"),None)
        saha=next((x for x in hareketler if x.durum=="sahaya_atandi"),None)
        bas=koord.tarih if koord else t.olusturulma_tarihi
        hedef=SLA_ATAMA_DAKIKA.get(t.oncelik,120)

        if saha and bas:
            dk=max(0,(saha.tarih-bas).total_seconds()/60)
            atama_sureleri.append(dk)
            if dk>hedef:
                sla_asan+=1
        elif t.durum in ("yeni","sefe_gonderildi"):
            canli=_sla_bilgisi(t)
            if canli["durum"]=="asildi":
                sla_asan+=1
            elif canli["durum"]=="yaklasiyor":
                sla_yaklasan+=1

    ort_atama_dk=(
        sum(atama_sureleri)/len(atama_sureleri)
        if atama_sureleri else None
    )

    # V39 — Behlül Bey'in istediği satır bazlı genel icmal:
    # İlçe + İş Türü + İş Alt Türü + Talep Eden Birim + Adet.
    icmal_gruplari={}
    for t in talepler:
        ilce=t.ilce.ad if t.ilce_id else "-"
        is_turu=t.is_turu.ad if t.is_turu_id else "-"
        is_alt_turu=t.is_alt_turu.ad if t.is_alt_turu_id else "-"
        acan_birim=_rapor_acan_birim(t)
        key=(ilce,is_turu,is_alt_turu,acan_birim)
        icmal_gruplari[key]=icmal_gruplari.get(key,0)+1

    satir_icmal=[
        {
            "ilce":key[0],
            "is_turu":key[1],
            "is_alt_turu":key[2],
            "acan_birim":key[3],
            "adet":adet,
        }
        for key,adet in icmal_gruplari.items()
    ]
    satir_icmal.sort(key=lambda x:(
        x["ilce"].casefold(),
        x["is_turu"].casefold(),
        x["is_alt_turu"].casefold(),
        x["acan_birim"].casefold(),
    ))

    def grup_ozeti(key_func,label_key):
        gruplar={}
        for e in is_emirleri:
            key=key_func(e) or "-"
            x=gruplar.setdefault(key,{
                label_key:key,"toplam":0,"acik":0,"tamam":0,"onay":0,"sureler":[]
            })
            x["toplam"]+=1
            if e.durum=="tamamlandi":
                x["tamam"]+=1
                if e.sef_onay_tarihi and e.atama_tarihi:
                    x["sureler"].append(
                        (e.sef_onay_tarihi-e.atama_tarihi).total_seconds()/60
                    )
            elif e.durum!="iptal":
                x["acik"]+=1
            if e.durum=="onay_bekliyor":
                x["onay"]+=1

        sonuc=[]
        for x in gruplar.values():
            ort=(sum(x["sureler"])/len(x["sureler"])) if x["sureler"] else None
            x["ortalama"]=_sure_etiketi(ort)
            x.pop("sureler",None)
            sonuc.append(x)
        return sorted(sonuc,key=lambda x:(-x["toplam"],str(x[label_key])))

    birim_icmal=grup_ozeti(
        lambda e:e.gonderen_birim or f"{e.talep.is_turu.ad} Birimi",
        "birim",
    )
    ilce_icmal=grup_ozeti(
        lambda e:e.talep.ilce.ad if e.talep.ilce_id else "-",
        "ilce",
    )
    is_turu_icmal=grup_ozeti(
        lambda e:e.talep.is_turu.ad if e.talep.is_turu_id else "-",
        "is_turu",
    )
    ekip_icmal=grup_ozeti(
        lambda e:(
            e.atanan_saha.kullanici.get_full_name()
            or e.atanan_saha.kullanici.username
        ) if e.atanan_saha_id else "Atanmamış",
        "ekip",
    )

    durumlar={}
    for e in is_emirleri:
        ad=e.get_durum_display()
        durumlar[ad]=durumlar.get(ad,0)+1
    durum_dagilimi=[
        {"durum":ad,"sayi":sayi}
        for ad,sayi in sorted(durumlar.items(),key=lambda x:(-x[1],x[0]))
    ]

    return {
        "talepler":talepler,
        "is_emirleri":is_emirleri,
        "kpi":{
            "toplam_talep":len(talepler),
            "toplam_is_emri":len(is_emirleri),
            "acik":len(acik_emir),
            "tamam":len(tamam_emir),
            "onay":sum(1 for e in is_emirleri if e.durum=="onay_bekliyor"),
            "atanmamis":sum(
                1 for t in talepler
                if not t.sorumlu_saha_id and t.durum not in ("tamamlandi","iptal")
            ),
            "acil_acik":sum(
                1 for e in acik_emir if e.talep.oncelik=="acil"
            ),
            "sla_asan":sla_asan,
            "sla_yaklasan":sla_yaklasan,
            "ortalama_atama":_sure_etiketi(ort_atama_dk),
            "ortalama_is_emri":_sure_etiketi(ort_is_emri_dk),
            "ortalama_cozum":_sure_etiketi(ort_cozum_dk),
        },
        "birim_icmal":birim_icmal,
        "ilce_icmal":ilce_icmal,
        "is_turu_icmal":is_turu_icmal,
        "ekip_icmal":ekip_icmal,
        "durum_dagilimi":durum_dagilimi,
        "satir_icmal":satir_icmal,
    }


def _abonelik_ambar_rapor_ozeti(baslangic, bitis):
    from abonelik.models import AmbarHareketi, AmbarSayacTalebi, Sozlesme, VatandasSicili
    return {
        "yeni_sicil": VatandasSicili.objects.filter(olusturulma_tarihi__date__range=(baslangic, bitis)).count(),
        "yeni_sozlesme": Sozlesme.objects.filter(olusturulma_tarihi__date__range=(baslangic, bitis)).count(),
        "ambar_talebi": AmbarSayacTalebi.objects.filter(talep_tarihi__date__range=(baslangic, bitis)).count(),
        "teslim": AmbarSayacTalebi.objects.filter(teslim_tarihi__date__range=(baslangic, bitis), durum="teslim_alindi").count(),
        "hurda": AmbarHareketi.objects.filter(tarih__date__range=(baslangic, bitis), islem="hurdaya_ayirma").count(),
    }


@panel_required("admin")
def sistem_rapor_merkezi(request):
    baslangic,bitis=_rapor_tarih_araligi(request)
    qs=_detayli_rapor_queryset(baslangic,bitis)
    icmal=_rapor_icmal_hazirla(baslangic,bitis)

    # Detay önizleme: ekran performansı için ilk 500 kayıt.
    onizleme=[]
    for t in qs[:500]:
        veri=_detayli_rapor_satiri(t)
        onizleme.append({
            "pk":t.pk,
            "talep_no":t.talep_no,
            "cells":[
                {"key":k,"label":label,"value":veri[k]}
                for k,label in DETAYLI_RAPOR_KOLONLARI
            ],
        })

    return render(request,"system/rapor_merkezi.html",{
        "kolonlar":[label for _,label in DETAYLI_RAPOR_KOLONLARI],
        "satirlar":onizleme,
        "toplam":len(icmal["talepler"]),
        "icmal":icmal["kpi"],
        "birim_icmal":icmal["birim_icmal"],
        "ilce_icmal":icmal["ilce_icmal"],
        "is_turu_icmal":icmal["is_turu_icmal"],
        "ekip_icmal":icmal["ekip_icmal"],
        "durum_dagilimi":icmal["durum_dagilimi"],
        "satir_icmal":icmal["satir_icmal"],
        "baslangic":baslangic.isoformat(),
        "bitis":bitis.isoformat(),
        "bugun":timezone.localdate().isoformat(),
        "bu_ay_baslangic":timezone.localdate().replace(day=1).isoformat(),
        "son_yedi_baslangic":(timezone.localdate()-timedelta(days=6)).isoformat(),
        "abonelik_ambar_rapor": _abonelik_ambar_rapor_ozeti(baslangic, bitis),
    })



@panel_required("admin")
@require_GET
def icmal_rapor_indir(request,format):
    baslangic,bitis=_rapor_tarih_araligi(request)
    data=_rapor_icmal_hazirla(baslangic,bitis)
    satirlar=data["satir_icmal"]
    format=(format or "").lower().strip()
    dosya_koku=f"isu185_genel_icmal_{baslangic.isoformat()}_{bitis.isoformat()}"
    rapor_tarihi=timezone.localtime().strftime("%d.%m.%Y %H:%M")
    logo_path=settings.BASE_DIR / "static" / "img" / "isu_report_logo.png"
    kolonlar=[
        ("ilce","İlçe"),
        ("is_turu","İş Türü"),
        ("is_alt_turu","İş Alt Türü"),
        ("acan_birim","Talep Eden Birim"),
        ("adet","Adet"),
    ]

    if format=="json":
        payload={
            "kurum":"Kocaeli Su ve Kanalizasyon İdaresi Genel Müdürlüğü",
            "sistem":"İSU 185 Talep Takip",
            "rapor":"Genel İcmal",
            "baslangic":baslangic.isoformat(),
            "bitis":bitis.isoformat(),
            "rapor_tarihi":rapor_tarihi,
            "toplam_talep":data["kpi"]["toplam_talep"],
            "satir_sayisi":len(satirlar),
            "kayitlar":[
                {label:satir[key] for key,label in kolonlar}
                for satir in satirlar
            ],
        }
        response=HttpResponse(
            json.dumps(payload,ensure_ascii=False,indent=2),
            content_type="application/json; charset=utf-8",
        )
        response["Content-Disposition"]=f'attachment; filename="{dosya_koku}.json"'
        return response

    if format=="csv":
        response=HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"]=f'attachment; filename="{dosya_koku}.csv"'
        response.write("\ufeff")
        wr=csv.writer(response,delimiter=";")
        wr.writerow(["KOCAELİ SU VE KANALİZASYON İDARESİ GENEL MÜDÜRLÜĞÜ"])
        wr.writerow(["İSU 185 TALEP TAKİP - GENEL İCMAL"])
        wr.writerow(["Dönem",baslangic.strftime("%d.%m.%Y"),bitis.strftime("%d.%m.%Y")])
        wr.writerow(["Rapor Tarihi",rapor_tarihi])
        wr.writerow([])
        wr.writerow([label for _key,label in kolonlar])
        for satir in satirlar:
            wr.writerow([satir[key] for key,_label in kolonlar])
        wr.writerow([])
        wr.writerow(["TOPLAM TALEP",data["kpi"]["toplam_talep"]])
        return response

    if format=="xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.drawing.image import Image as XLImage
        except ImportError:
            return HttpResponse(
                "Excel dışa aktarımı için openpyxl kurulmalıdır.",
                status=500,
                content_type="text/plain; charset=utf-8",
            )

        wb=Workbook()
        ws=wb.active
        ws.title="Genel İcmal"
        ws.sheet_view.showGridLines=False

        green="1E5A42"
        pale="E8F0E5"
        stripe="F8FAF8"
        thin=Side(style="thin",color="C8D4CB")

        if logo_path.exists():
            logo=XLImage(str(logo_path))
            logo.width=235
            logo.height=60
            ws.add_image(logo,"A1")

        ws.merge_cells("D1:H1")
        ws["D1"]="İSU 185 TALEP TAKİP - GENEL İCMAL"
        ws["D1"].font=Font(bold=True,size=16,color=green)
        ws["D2"]=f"Dönem: {baslangic.strftime('%d.%m.%Y')} - {bitis.strftime('%d.%m.%Y')}"
        ws["D2"].font=Font(size=9,color="55665D")
        ws["D3"]=f"Toplam Talep: {data['kpi']['toplam_talep']}"
        ws["D3"].font=Font(bold=True,size=9,color=green)
        ws["H2"]="Rapor Tarihi"
        ws["I2"]=rapor_tarihi
        ws["H2"].font=Font(bold=True,size=8,color="5E6C64")
        ws["I2"].font=Font(size=8,color="263C32")

        header_row=5
        for idx,(_key,label) in enumerate(kolonlar,1):
            c=ws.cell(header_row,idx,label)
            c.font=Font(bold=True,size=9,color="334A3E")
            c.fill=PatternFill("solid",fgColor=pale)
            c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            c.border=Border(top=thin,bottom=thin,left=thin,right=thin)

        for r_idx,satir in enumerate(satirlar,header_row+1):
            for c_idx,(key,_label) in enumerate(kolonlar,1):
                c=ws.cell(r_idx,c_idx,satir[key])
                c.font=Font(size=9,color="263C32")
                c.fill=PatternFill("solid",fgColor="FFFFFF" if r_idx%2 else stripe)
                c.border=Border(top=thin,bottom=thin,left=thin,right=thin)
                c.alignment=Alignment(
                    horizontal="center" if key=="adet" else "left",
                    vertical="center",wrap_text=True,
                )

        ws.column_dimensions["A"].width=22
        ws.column_dimensions["B"].width=30
        ws.column_dimensions["C"].width=36
        ws.column_dimensions["D"].width=28
        ws.column_dimensions["E"].width=12
        ws.freeze_panes="A6"
        ws.auto_filter.ref=f"A5:E{max(5,ws.max_row)}"
        ws.row_dimensions[5].height=27
        ws.page_setup.orientation="landscape"
        ws.sheet_properties.pageSetUpPr.fitToPage=True
        ws.page_setup.fitToWidth=1
        ws.page_setup.fitToHeight=0

        output=io.BytesIO()
        wb.save(output)
        response=HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"]=f'attachment; filename="{dosya_koku}.xlsx"'
        return response

    if format=="pdf":
        try:
            import html
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.platypus import SimpleDocTemplate, LongTable, TableStyle, Paragraph
        except ImportError:
            return HttpResponse(
                "PDF dışa aktarımı için reportlab kurulmalıdır.",
                status=500,
                content_type="text/plain; charset=utf-8",
            )

        font_name="Helvetica"
        for font_path in (
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/calibri.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ):
            try:
                pdfmetrics.registerFont(TTFont("ISUV39IcmalFont",font_path))
                font_name="ISUV39IcmalFont"
                break
            except Exception:
                continue

        page_size=landscape(A4)
        output=io.BytesIO()
        doc=SimpleDocTemplate(
            output,pagesize=page_size,
            leftMargin=10*mm,rightMargin=10*mm,topMargin=30*mm,bottomMargin=11*mm,
            title="İSU 185 Genel İcmal Raporu",author="İSU Genel Müdürlüğü",
        )

        def header_footer(canvas,doc_obj):
            w,h=page_size
            canvas.saveState()
            if logo_path.exists():
                try:
                    canvas.drawImage(str(logo_path),10*mm,h-24*mm,width=55*mm,height=14*mm,preserveAspectRatio=True,mask="auto")
                except Exception:
                    pass
            canvas.setFont(font_name,12)
            canvas.setFillColor(colors.HexColor("#174C37"))
            canvas.drawCentredString(w/2,h-11*mm,"İSU 185 TALEP TAKİP SİSTEMİ")
            canvas.setFont(font_name,8)
            canvas.setFillColor(colors.HexColor("#50645A"))
            canvas.drawCentredString(w/2,h-16*mm,"GENEL İCMAL RAPORU")
            canvas.setFont(font_name,6.5)
            canvas.drawRightString(w-10*mm,h-10*mm,f"Dönem: {baslangic.strftime('%d.%m.%Y')} - {bitis.strftime('%d.%m.%Y')}")
            canvas.drawRightString(w-10*mm,h-14*mm,f"Toplam Talep: {data['kpi']['toplam_talep']}")
            canvas.drawRightString(w-10*mm,h-18*mm,f"Sayfa: {doc_obj.page}")
            canvas.setStrokeColor(colors.HexColor("#1F6A4A"))
            canvas.line(10*mm,h-26*mm,w-10*mm,h-26*mm)
            canvas.setFont(font_name,5.5)
            canvas.setFillColor(colors.HexColor("#6E7D75"))
            canvas.drawString(10*mm,5*mm,"Kocaeli Su ve Kanalizasyon İdaresi Genel Müdürlüğü")
            canvas.drawRightString(w-10*mm,5*mm,f"Rapor Tarihi: {rapor_tarihi}")
            canvas.restoreState()

        styles=getSampleStyleSheet()
        cell=ParagraphStyle(
            "V39IcmalCell",parent=styles["Normal"],fontName=font_name,
            fontSize=7.3,leading=8.5,textColor=colors.HexColor("#263C32"),
        )
        hcell=ParagraphStyle(
            "V39IcmalHeader",parent=cell,fontSize=7.1,leading=8.3,
            textColor=colors.HexColor("#334A3E"),
        )
        table_data=[
            [Paragraph(f"<b>{html.escape(label)}</b>",hcell) for _key,label in kolonlar]
        ]
        for satir in satirlar:
            table_data.append([
                Paragraph(html.escape(str(satir[key])),cell) if key!="adet" else str(satir[key])
                for key,_label in kolonlar
            ])
        if not satirlar:
            table_data.append([Paragraph("Kayıt bulunamadı.",cell),"","","",""])

        tb=LongTable(
            table_data,
            colWidths=[38*mm,53*mm,67*mm,48*mm,20*mm],
            repeatRows=1,hAlign="CENTER",
        )
        tb.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#DFEBDD")),
            ("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#CBD6CE")),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F8FAF8")]),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("ALIGN",(-1,1),(-1,-1),"CENTER"),
            ("LEFTPADDING",(0,0),(-1,-1),4),
            ("RIGHTPADDING",(0,0),(-1,-1),4),
            ("TOPPADDING",(0,0),(-1,-1),4),
            ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ]))
        doc.build([tb],onFirstPage=header_footer,onLaterPages=header_footer)
        response=HttpResponse(output.getvalue(),content_type="application/pdf")
        response["Content-Disposition"]=f'attachment; filename="{dosya_koku}.pdf"'
        return response

    return HttpResponse(
        "Desteklenmeyen icmal formatı. PDF, XLSX, CSV veya JSON kullanın.",
        status=400,
        content_type="text/plain; charset=utf-8",
    )


@panel_required("admin")
@require_GET
def detayli_rapor_indir(request,format):
    baslangic,bitis=_rapor_tarih_araligi(request)
    qs=_detayli_rapor_queryset(baslangic,bitis)
    satirlar=[_detayli_rapor_satiri(t) for t in qs]
    format=(format or "").lower().strip()
    dosya_koku=f"isu185_detayli_rapor_{baslangic.isoformat()}_{bitis.isoformat()}"
    rapor_tarihi=timezone.localtime().strftime("%d.%m.%Y %H:%M")
    logo_path=settings.BASE_DIR / "static" / "img" / "isu_report_logo.png"

    if format=="csv":
        response=HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"]=f'attachment; filename="{dosya_koku}.csv"'
        response.write("\ufeff")
        wr=csv.writer(response,delimiter=";")
        wr.writerow([label for _,label in DETAYLI_RAPOR_KOLONLARI])
        for satir in satirlar:
            wr.writerow([satir[k] for k,_ in DETAYLI_RAPOR_KOLONLARI])
        return response

    if format=="json":
        payload={
            "kurum":"Kocaeli Su ve Kanalizasyon İdaresi Genel Müdürlüğü",
            "sistem":"İSU 185 Talep Takip",
            "rapor":"Detaylı Talep / İş Emri Raporu",
            "baslangic":baslangic.isoformat(),
            "bitis":bitis.isoformat(),
            "rapor_tarihi":rapor_tarihi,
            "kayit_sayisi":len(satirlar),
            "kayitlar":[
                {label:satir[k] for k,label in DETAYLI_RAPOR_KOLONLARI}
                for satir in satirlar
            ],
        }
        response=HttpResponse(
            json.dumps(payload,ensure_ascii=False,indent=2),
            content_type="application/json; charset=utf-8",
        )
        response["Content-Disposition"]=f'attachment; filename="{dosya_koku}.json"'
        return response

    if format=="xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.drawing.image import Image as XLImage
        except ImportError:
            return HttpResponse(
                "Excel dışa aktarımı için openpyxl kurulmalıdır.",
                status=500,
                content_type="text/plain; charset=utf-8",
            )

        wb=Workbook()
        ws=wb.active
        ws.title="Detaylı Rapor"
        ws.sheet_view.showGridLines=False

        green="1E5A42"
        pale="DFEBDD"
        stripe="F8FAF8"
        border_color="CCD7CF"
        thin=Side(style="thin",color=border_color)

        if logo_path.exists():
            logo=XLImage(str(logo_path))
            logo.width=235
            logo.height=60
            ws.add_image(logo,"A1")
        ws.merge_cells("D1:J1")
        ws["D1"]="İSU 185 TALEP TAKİP - DETAYLI TALEP / İŞ EMRİ RAPORU"
        ws["D1"].font=Font(bold=True,size=15,color=green)
        ws["D1"].alignment=Alignment(vertical="center")
        ws.merge_cells("D2:J2")
        ws["D2"]="1 satır = 1 talebin baştan sona operasyon kaydı"
        ws["D2"].font=Font(size=9,color="5E6C64")
        ws["M1"]="Dönem"
        ws["N1"]=f"{baslangic.strftime('%d.%m.%Y')} - {bitis.strftime('%d.%m.%Y')}"
        ws["M2"]="Rapor Tarihi"
        ws["N2"]=rapor_tarihi
        ws["M3"]="Toplam Kayıt"
        ws["N3"]=len(satirlar)
        for c in ("M1","M2","M3"):
            ws[c].font=Font(bold=True,size=8,color="5E6C64")
        for c in ("N1","N2","N3"):
            ws[c].font=Font(size=8,color="263C32")

        header_row=5
        headers=[label for _,label in DETAYLI_RAPOR_KOLONLARI]
        for idx,label in enumerate(headers,1):
            cell=ws.cell(header_row,idx,label)
            cell.font=Font(bold=True,size=7,color="334A3E")
            cell.fill=PatternFill("solid",fgColor=pale)
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            cell.border=Border(top=thin,bottom=thin,left=thin,right=thin)

        durum_idx=next(i for i,(k,_) in enumerate(DETAYLI_RAPOR_KOLONLARI,1) if k=="durum")
        sla_idx=next(i for i,(k,_) in enumerate(DETAYLI_RAPOR_KOLONLARI,1) if k=="sla_atama_sonucu")
        for r_idx,satir in enumerate(satirlar,header_row+1):
            for c_idx,(key,_label) in enumerate(DETAYLI_RAPOR_KOLONLARI,1):
                value=satir[key]
                cell=ws.cell(r_idx,c_idx,value)
                cell.font=Font(size=7,color="263C32")
                cell.alignment=Alignment(vertical="top",wrap_text=True)
                cell.fill=PatternFill("solid",fgColor="FFFFFF" if r_idx%2 else stripe)
                cell.border=Border(top=thin,bottom=thin,left=thin,right=thin)
            durum=str(ws.cell(r_idx,durum_idx).value or "")
            sla=str(ws.cell(r_idx,sla_idx).value or "")
            if "Tamam" in durum:
                ws.cell(r_idx,durum_idx).font=Font(size=7,bold=True,color="2E7A3F")
            elif "Onay" in durum:
                ws.cell(r_idx,durum_idx).font=Font(size=7,bold=True,color="A06B13")
            elif "Açık" in durum or "Yeni" in durum:
                ws.cell(r_idx,durum_idx).font=Font(size=7,bold=True,color="A14138")
            if "AŞILDI" in sla.upper() or "Aşıldı" in sla:
                ws.cell(r_idx,sla_idx).font=Font(size=7,bold=True,color="A14138")
            elif "Zamanında" in sla or "Süre içinde" in sla:
                ws.cell(r_idx,sla_idx).font=Font(size=7,bold=True,color="2E7A3F")

        width_by_key={
            "talep_no":18,"is_emri_no":18,"talep_baslangic":19,"talep_adi":30,
            "vatandas":22,"telefon":16,"eposta":28,"ilce":15,"mahalle":23,"yol":28,
            "kapi_no":10,"adres":42,"koordinat":22,"is_turu":24,"is_alt_turu":28,
            "oncelik":12,"durum":20,"gonderen_birim":30,"olusturan":20,
            "koordinator":28,"koordinator_rol":24,"koordinator_tarihi":20,
            "saha":27,"saha_rol":28,"saha_atama_tarihi":20,"sla_atama_suresi":22,
            "saha_kabul_tarihi":20,"yola_cikis_tarihi":20,"adrese_ulasma_tarihi":20,
            "mudahale_baslama_tarihi":20,"saha_bitirdi":20,"sef_onaylayan":22,
            "sef_onay_tarihi":20,"tekrar_baslatildi":16,"tekrar_sayisi":14,
            "tamamlanma_tarihi":20,"is_emri_suresi":20,"toplam_cozum_suresi":22,
            "sla_atama_hedefi":16,"sla_atama_sonucu":30,"gps_dogrulama":17,
            "gps_mesafe":14,"zorunlu_gorsel":16,"gorsel_durumu":20,"geri_donus_durumu":22,
            "geri_donus_tarihi":20,"geri_donus_personeli":23,"gorusme_sonucu":22,
            "genel_memnuniyet":18,"anket_puan":14,"talep_aciklama":46,
            "saha_sonuc":46,"gorusme_notu":42,
        }
        for idx,(key,_label) in enumerate(DETAYLI_RAPOR_KOLONLARI,1):
            ws.column_dimensions[ws.cell(1,idx).column_letter].width=width_by_key.get(key,18)

        ws.freeze_panes="C6"
        ws.auto_filter.ref=f"A5:{ws.cell(ws.max_row,ws.max_column).coordinate}"
        ws.row_dimensions[5].height=34
        ws.sheet_properties.pageSetUpPr.fitToPage=False
        ws.page_setup.orientation="landscape"

        # V39 — Detaylı Excel içindeki Genel İcmal sayfası da resmi satır bazlı yapıdadır.
        icmal=_rapor_icmal_hazirla(baslangic,bitis)
        summary=wb.create_sheet("Genel İcmal",0)
        summary.sheet_view.showGridLines=False
        if logo_path.exists():
            logo2=XLImage(str(logo_path))
            logo2.width=235
            logo2.height=60
            summary.add_image(logo2,"A1")
        summary["D1"]="İSU 185 — GENEL İCMAL"
        summary["D1"].font=Font(bold=True,size=15,color=green)
        summary["D2"]=f"Dönem: {baslangic.strftime('%d.%m.%Y')} - {bitis.strftime('%d.%m.%Y')}"
        summary["D2"].font=Font(size=9,color="5E6C64")
        summary["D3"]=f"Toplam Talep: {icmal['kpi']['toplam_talep']}"
        summary["D3"].font=Font(bold=True,size=9,color=green)

        icmal_cols=[
            ("ilce","İlçe"),("is_turu","İş Türü"),("is_alt_turu","İş Alt Türü"),
            ("acan_birim","Talep Eden Birim"),("adet","Adet"),
        ]
        for idx,(_key,label) in enumerate(icmal_cols,1):
            c=summary.cell(5,idx,label)
            c.fill=PatternFill("solid",fgColor=pale)
            c.font=Font(bold=True,size=9,color="405249")
            c.border=Border(top=thin,bottom=thin,left=thin,right=thin)
            c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        for r_idx,row in enumerate(icmal["satir_icmal"],6):
            for c_idx,(key,_label) in enumerate(icmal_cols,1):
                c=summary.cell(r_idx,c_idx,row[key])
                c.border=Border(top=thin,bottom=thin,left=thin,right=thin)
                c.fill=PatternFill("solid",fgColor="FFFFFF" if r_idx%2 else stripe)
                c.font=Font(size=9,color="263C32")
                c.alignment=Alignment(horizontal="center" if key=="adet" else "left",vertical="center",wrap_text=True)
        for col,width in {"A":22,"B":30,"C":36,"D":28,"E":12}.items():
            summary.column_dimensions[col].width=width
        summary.freeze_panes="A6"
        summary.auto_filter.ref=f"A5:E{max(5,summary.max_row)}"
        summary.page_setup.orientation="landscape"

        output=io.BytesIO()
        wb.save(output)
        response=HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"]=f'attachment; filename="{dosya_koku}.xlsx"'
        return response

    if format=="pdf":
        try:
            import html
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.platypus import SimpleDocTemplate, LongTable, TableStyle, Paragraph
        except ImportError:
            return HttpResponse(
                "PDF dışa aktarımı için reportlab kurulmalıdır.",
                status=500,
                content_type="text/plain; charset=utf-8",
            )

        font_name="Helvetica"
        for font_path in (
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/calibri.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ):
            try:
                pdfmetrics.registerFont(TTFont("ISUV33DetailFont",font_path))
                font_name="ISUV33DetailFont"
                break
            except Exception:
                continue

        widths_mm={
            "talep_no":28,"is_emri_no":28,"talep_baslangic":30,"talep_adi":44,
            "vatandas":36,"telefon":27,"eposta":42,"ilce":23,"mahalle":34,"yol":42,
            "kapi_no":17,"adres":68,"koordinat":36,"is_turu":39,"is_alt_turu":42,
            "oncelik":20,"durum":30,"talep_eden_birim":38,"gonderen_birim":46,"olusturan":34,
            "koordinator":45,"koordinator_rol":38,"koordinator_tarihi":31,
            "saha":44,"saha_rol":43,"saha_atama_tarihi":31,"sla_atama_suresi":34,
            "saha_kabul_tarihi":31,"yola_cikis_tarihi":31,"adrese_ulasma_tarihi":31,
            "mudahale_baslama_tarihi":31,"saha_bitirdi":31,"sef_onaylayan":37,
            "sef_onay_tarihi":31,"tekrar_baslatildi":28,"tekrar_sayisi":24,
            "tamamlanma_tarihi":31,"is_emri_suresi":32,"toplam_cozum_suresi":34,
            "sla_atama_hedefi":27,"sla_atama_sonucu":43,"gps_dogrulama":27,
            "gps_mesafe":24,"zorunlu_gorsel":28,"gorsel_durumu":32,"geri_donus_durumu":36,
            "geri_donus_tarihi":31,"geri_donus_personeli":38,"gorusme_sonucu":36,
            "genel_memnuniyet":30,"anket_puan":24,"talep_aciklama":70,
            "saha_sonuc":70,"gorusme_notu":65,
        }
        raw_widths=[widths_mm.get(key,30) for key,_ in DETAYLI_RAPOR_KOLONLARI]
        # PDF tek geniş yatay sayfa mantığında kalır; ancak okunabilirliği korumak
        # için toplam tablo genişliği yaklaşık 1450 mm ile sınırlandırılır.
        # Böylece PDF görüntüleyicide normal zoomda metin okunur, yatay kaydırma yapılır.
        scale=min(1.0,1450.0/sum(raw_widths))
        col_widths=[w*scale*mm for w in raw_widths]
        page_width=sum(col_widths)+(16*mm)
        page_height=210*mm
        page_size=(page_width,page_height)

        output=io.BytesIO()
        doc=SimpleDocTemplate(
            output,pagesize=page_size,
            leftMargin=8*mm,rightMargin=8*mm,topMargin=29*mm,bottomMargin=10*mm,
            title="İSU 185 Detaylı Talep İş Emri Raporu",author="İSU Genel Müdürlüğü",
        )

        def header_footer(canvas,doc_obj):
            w,h=page_size
            canvas.saveState()
            if logo_path.exists():
                try:
                    canvas.drawImage(str(logo_path),8*mm,h-24*mm,width=58*mm,height=15*mm,preserveAspectRatio=True,mask="auto")
                except Exception:
                    pass
            canvas.setFont(font_name,11)
            canvas.setFillColor(colors.HexColor("#174C37"))
            canvas.drawString(72*mm,h-11*mm,"İSU 185 TALEP TAKİP SİSTEMİ")
            canvas.setFont(font_name,7)
            canvas.setFillColor(colors.HexColor("#50645A"))
            canvas.drawString(72*mm,h-16*mm,"DETAYLI TALEP / İŞ EMRİ RAPORU - 1 SATIR = 1 TALEP")
            canvas.setFont(font_name,5.8)
            canvas.drawString(280*mm,h-10*mm,f"Dönem: {baslangic.strftime('%d.%m.%Y')} - {bitis.strftime('%d.%m.%Y')}")
            canvas.drawString(280*mm,h-14*mm,f"Rapor Tarihi: {rapor_tarihi}")
            canvas.drawString(280*mm,h-18*mm,f"Kayıt: {len(satirlar)} • Sayfa: {doc_obj.page}")
            canvas.setStrokeColor(colors.HexColor("#1F6A4A"))
            canvas.setLineWidth(0.7)
            canvas.line(8*mm,h-26*mm,w-8*mm,h-26*mm)
            canvas.setFont(font_name,5)
            canvas.setFillColor(colors.HexColor("#6E7D75"))
            canvas.drawString(8*mm,4*mm,"Kocaeli Su ve Kanalizasyon İdaresi Genel Müdürlüğü • ALO 185")
            canvas.drawString(160*mm,4*mm,"Veriler İSU 185 Talep Takip Sistemi tarafından otomatik oluşturulmuştur.")
            canvas.restoreState()

        styles=getSampleStyleSheet()
        header_style=ParagraphStyle(
            "V33DetailHeader",parent=styles["Normal"],fontName=font_name,
            fontSize=6.0,leading=7.0,textColor=colors.HexColor("#32483C"),
        )
        cell_style=ParagraphStyle(
            "V33DetailCell",parent=styles["Normal"],fontName=font_name,
            fontSize=5.8,leading=6.8,textColor=colors.HexColor("#263C32"),
        )

        table_data=[
            [Paragraph(f"<b>{html.escape(label)}</b>",header_style) for _,label in DETAYLI_RAPOR_KOLONLARI]
        ]
        for satir in satirlar:
            row=[]
            for key,_label in DETAYLI_RAPOR_KOLONLARI:
                raw=satir.get(key,"-")
                value="-" if raw is None or raw=="" else str(raw)
                row.append(Paragraph(html.escape(value),cell_style))
            table_data.append(row)

        if not satirlar:
            empty=["" for _ in DETAYLI_RAPOR_KOLONLARI]
            empty[0]="Seçilen dönemde kayıt bulunamadı."
            table_data.append([Paragraph(html.escape(x),cell_style) for x in empty])

        table=LongTable(table_data,colWidths=col_widths,repeatRows=1,hAlign="LEFT")
        style=[
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#DFEBDD")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.HexColor("#32483C")),
            ("GRID",(0,0),(-1,-1),0.18,colors.HexColor("#CBD6CE")),
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("LEFTPADDING",(0,0),(-1,-1),1.6),
            ("RIGHTPADDING",(0,0),(-1,-1),1.6),
            ("TOPPADDING",(0,0),(-1,-1),2.6),
            ("BOTTOMPADDING",(0,0),(-1,-1),2.6),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F8FAF8")]),
            ("BACKGROUND",(0,1),(1,-1),colors.HexColor("#F0F6ED")),
        ]
        durum_col=next(i for i,(k,_) in enumerate(DETAYLI_RAPOR_KOLONLARI) if k=="durum")
        sla_col=next(i for i,(k,_) in enumerate(DETAYLI_RAPOR_KOLONLARI) if k=="sla_atama_sonucu")
        for ridx,satir in enumerate(satirlar,1):
            durum=str(satir.get("durum",""))
            sla=str(satir.get("sla_atama_sonucu",""))
            if "Tamam" in durum:
                style.append(("TEXTCOLOR",(durum_col,ridx),(durum_col,ridx),colors.HexColor("#2E7A3F")))
            elif "Onay" in durum:
                style.append(("TEXTCOLOR",(durum_col,ridx),(durum_col,ridx),colors.HexColor("#A06B13")))
            elif "Yeni" in durum or "Açık" in durum:
                style.append(("TEXTCOLOR",(durum_col,ridx),(durum_col,ridx),colors.HexColor("#A14138")))
            if "AŞILDI" in sla.upper() or "Aşıldı" in sla:
                style.append(("TEXTCOLOR",(sla_col,ridx),(sla_col,ridx),colors.HexColor("#A14138")))
            elif "Zamanında" in sla or "Süre içinde" in sla:
                style.append(("TEXTCOLOR",(sla_col,ridx),(sla_col,ridx),colors.HexColor("#2E7A3F")))
        table.setStyle(TableStyle(style))

        doc.build([table],onFirstPage=header_footer,onLaterPages=header_footer)
        response=HttpResponse(output.getvalue(),content_type="application/pdf")
        response["Content-Disposition"]=f'attachment; filename="{dosya_koku}_genis_yatay.pdf"'
        return response

    return HttpResponse(
        "Desteklenmeyen rapor formatı. PDF, XLSX, CSV veya JSON seçin.",
        status=400,
        content_type="text/plain; charset=utf-8",
    )


@panel_required("admin")
def sistem_raporlar(request):
    return render(
        request,
        "system/raporlar.html",
        _otomatik_rapor_context(),
    )


@panel_required("admin")
@require_GET
def sistem_rapor_canli(request):
    """
    Rapor ekranı bunu periyodik kontrol eder.
    Yeni talep veya herhangi bir durum/log değişikliği olduğunda
    ekran kendini otomatik yeniler.
    """
    ctx=_otomatik_rapor_context()
    return JsonResponse({
        "version":ctx["rapor_version"],
        "checked_at":timezone.localtime().strftime("%H:%M:%S"),
        "toplam":ctx["toplam"],
        "yeni":ctx["yeni"],
        "sahada":ctx["sahada"],
        "onay":ctx["onay"],
        "tamam":ctx["tamam"],
        "geri_bekleyen":ctx["geri_bekleyen"],
    })


@panel_required("admin")
def rapor_csv(request):
    qs=Talep.objects.select_related(
        "ilce","mahalle","yol","is_turu","is_alt_turu"
    ).order_by("-olusturulma_tarihi")
    response=HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"]='attachment; filename="isu185_otomatik_is_listesi.csv"'
    response.write("\ufeff")
    wr=csv.writer(response)
    wr.writerow([
        "Talep No","Tarih","İlçe","Mahalle","Cadde/Sokak",
        "İş Türü","Alt Tür","Öncelik","Durum",
        "Vatandaş Geri Dönüş","Şef Onay Tarihi"
    ])
    for t in qs:
        wr.writerow([
            t.talep_no,
            timezone.localtime(t.olusturulma_tarihi).strftime("%d.%m.%Y %H:%M"),
            t.ilce.ad,t.mahalle.ad,t.yol.ad,
            t.is_turu.ad,t.is_alt_turu.ad,
            t.get_oncelik_display(),t.get_durum_display(),
            t.get_vatandas_bildirim_durumu_display(),
            timezone.localtime(t.sef_onay_tarihi).strftime("%d.%m.%Y %H:%M") if t.sef_onay_tarihi else "",
        ])
    return response



@panel_required("admin")
def sistem_operasyon_merkezi(request):
    aktif_emirler=list(
        IsEmri.objects.exclude(durum__in=["tamamlandi","iptal"]).select_related(
            "talep__ilce","talep__mahalle","talep__is_turu","talep__is_alt_turu",
            "atanan_saha__kullanici","atanan_saha__rol",
        ).order_by("talep__oncelik","atama_tarihi")
    )

    saha_profilleri=list(
        PersonelProfili.objects.filter(
            rol__panel_tipi="saha",
            aktif=True,
            kullanici__is_active=True,
        ).select_related("kullanici","rol").prefetch_related(
            "yetkili_ilceler","uzmanlik_is_turleri"
        )
    )
    ekipler=[]
    for p in saha_profilleri:
        acik=[e for e in aktif_emirler if e.atanan_saha_id==p.id]
        ekipler.append({
            "profil":p,
            "acik":len(acik),
            "acil":sum(1 for e in acik if e.talep.oncelik=="acil"),
            "son_is":acik[0] if acik else None,
        })
    ekipler.sort(key=lambda x:(-x["acil"],-x["acik"],x["profil"].rol.ad,x["profil"].kullanici.username))

    # Henüz sahaya atanmamış işlerin SLA kontrolü.
    bekleyen=list(
        Talep.objects.filter(durum__in=["yeni","sefe_gonderildi"]).select_related(
            "ilce","mahalle","is_turu","is_alt_turu","sorumlu_koordinator__kullanici"
        )
    )
    geciken=[]
    yaklasan=[]
    for t in bekleyen:
        sla=_sla_bilgisi(t)
        row={"talep":t,"sla":sla}
        if sla["durum"]=="asildi":
            geciken.append(row)
        elif sla["durum"]=="yaklasiyor":
            yaklasan.append(row)
    geciken.sort(key=lambda x:-x["sla"]["asma_dakika"])
    yaklasan.sort(key=lambda x:x["sla"]["kalan_dakika"])

    ilce_map={}
    for e in aktif_emirler:
        ad=e.talep.ilce.ad
        x=ilce_map.setdefault(ad,{"ilce":ad,"aktif":0,"acil":0})
        x["aktif"]+=1
        if e.talep.oncelik=="acil":
            x["acil"]+=1
    ilce_yogunluk=sorted(ilce_map.values(),key=lambda x:(-x["aktif"],x["ilce"]))

    harita=[]
    for e in aktif_emirler:
        t=e.talep
        # V37 — Operasyon Merkezi de bozuk/ülke dışına taşan koordinatlar
        # yüzünden tüm haritayı uzaklaştırmasın.
        if not _kocaeli_harita_koordinati_gecerli(t.lat,t.lng):
            continue
        harita.append({
            "lat":float(t.lat),
            "lng":float(t.lng),
            "is_emri":e.is_emri_no,
            "talep":t.talep_no,
            "ilce":t.ilce.ad,
            "is_turu":t.is_turu.ad,
            "oncelik":t.get_oncelik_display(),
            "oncelik_kod":t.oncelik,
            "durum":e.get_durum_display(),
            "ekip":(
                e.atanan_saha.kullanici.get_full_name()
                or e.atanan_saha.kullanici.username
            ) if e.atanan_saha_id else "Atanmamış",
            "url":reverse("dashboard:is_emri_detay",args=[e.pk]),
        })

    return render(request,"system/operasyon_merkezi.html",{
        "aktif_emirler":aktif_emirler[:150],
        "ekipler":ekipler,
        "geciken":geciken[:50],
        "yaklasan":yaklasan[:50],
        "ilce_yogunluk":ilce_yogunluk,
        "harita_json":harita,
        "kpi":{
            "aktif":len(aktif_emirler),
            "acil":sum(1 for e in aktif_emirler if e.talep.oncelik=="acil"),
            "sahada":sum(1 for e in aktif_emirler if e.durum in ("yolda","yerinde","islemde")),
            "onay":sum(1 for e in aktif_emirler if e.durum=="onay_bekliyor"),
            "musait":sum(1 for p in saha_profilleri if p.musait),
            "geciken":len(geciken),
        },
    })


# =====================================================================
# V28 — ABONE / İŞ EMRİ / MOBİL SAHA
# =====================================================================

@panel_required("admin")
def sistem_aboneler(request):
    q=(request.GET.get("q") or "").strip()
    if request.method=="POST":
        form=AboneForm(request.POST)
        if form.is_valid():
            abone=form.save()
            messages.success(request,f"{abone.abone_no} numaralı abone kaydedildi.")
            return redirect("dashboard:sistem_aboneler")
    else:
        form=AboneForm()

    qs=Abone.objects.select_related("ilce","mahalle","yol").order_by("abone_no")
    if q:
        qs=qs.filter(
            Q(abone_no__icontains=q)|
            Q(ad__icontains=q)|
            Q(soyad__icontains=q)|
            Q(telefon__icontains=q)|
            Q(sayac_no__icontains=q)
        )
    return render(request,"system/aboneler.html",{
        "form":form,
        "aboneler":qs[:300],
        "toplam":Abone.objects.count(),
        "aktif":Abone.objects.filter(aktif=True).count(),
        "q":q,
    })


@panel_required("185","admin")
@require_GET
def abone_sorgula(request):
    no=(request.GET.get("no") or "").strip()
    if not no:
        return JsonResponse({"ok":False,"message":"Abone numarası girin."},status=400)
    abone=Abone.objects.select_related("ilce","mahalle","yol").filter(
        abone_no__iexact=no,aktif=True
    ).first()
    if not abone:
        return JsonResponse({"ok":False,"message":"Aktif abone kaydı bulunamadı."},status=404)
    return JsonResponse({
        "ok":True,
        "id":abone.pk,
        "abone_no":abone.abone_no,
        "ad":abone.ad,
        "soyad":abone.soyad,
        "telefon":abone.telefon,
        "eposta":abone.eposta,
        "sayac_no":abone.sayac_no,
        "adres":abone.tam_adres,
        "ilce":abone.ilce.ad if abone.ilce_id else "",
        "mahalle":abone.mahalle.ad if abone.mahalle_id else "",
        "yol":abone.yol.ad if abone.yol_id else "",
        "kapi_no":abone.kapi_no,
    })


@panel_required("admin")
def sistem_is_emirleri(request):
    durum=(request.GET.get("durum") or "").strip()
    q=(request.GET.get("q") or "").strip()
    qs=IsEmri.objects.select_related(
        "talep__ilce","talep__mahalle","talep__is_turu","talep__is_alt_turu",
        "atanan_saha__kullanici","atanan_saha__rol","olusturan"
    )
    if durum:
        qs=qs.filter(durum=durum)
    if q:
        qs=qs.filter(
            Q(is_emri_no__icontains=q)|
            Q(talep__talep_no__icontains=q)|
            Q(gonderen_birim__icontains=q)|
            Q(talep__vatandas_ad__icontains=q)|
            Q(talep__vatandas_soyad__icontains=q)
        )

    tum=IsEmri.objects.all()
    tamam=tum.filter(durum="tamamlandi")
    sureler=[
        (x.sef_onay_tarihi-x.atama_tarihi).total_seconds()/60
        for x in tamam if x.sef_onay_tarihi and x.atama_tarihi
    ]
    ort_dk=(sum(sureler)/len(sureler)) if sureler else 0
    ortalama=f"{int(ort_dk//60)} sa {int(ort_dk%60)} dk" if ort_dk>=60 else (f"{int(ort_dk)} dk" if ort_dk else "-")
    return render(request,"system/is_emirleri.html",{
        "is_emirleri":qs[:500],
        "toplam":tum.count(),
        "acik":tum.exclude(durum__in=["tamamlandi","iptal"]).count(),
        "tamam":tamam.count(),
        "onay":tum.filter(durum="onay_bekliyor").count(),
        "ortalama":ortalama,
        "durum":durum,
        "q":q,
        "durumlar":IsEmri.DURUMLAR,
    })


@login_required
@require_POST
def is_emri_fotograf_yukle(request,pk):
    emri=get_object_or_404(
        IsEmri.objects.select_related("talep__is_alt_turu","atanan_saha__kullanici"),pk=pk
    )
    if not request.user.is_superuser:
        p=get_profile(request.user)
        if not p or p.rol.panel_tipi!="saha" or emri.atanan_saha_id!=p.id:
            raise PermissionDenied

    p=get_profile(request.user) if not request.user.is_superuser else None
    mobil=request.POST.get("next")=="mobil"
    if p and _web_only_abone_role(p.rol):
        mobil=False
        hedef="dashboard:abone_gorevleri"
    else:
        hedef="dashboard:mobil_saha" if mobil else "dashboard:saha"
    if emri.durum not in {"yerinde","islemde"}:
        messages.error(request,"İş emri görselleri adrese ulaşıldıktan sonra eklenebilir.")
        return redirect(hedef)

    try:
        sira=int(request.POST.get("sira") or 0)
    except ValueError:
        sira=0
    etiketler=emri.talep.is_alt_turu.zorunlu_foto_etiketleri
    if not 1<=sira<=len(etiketler):
        messages.error(request,"Geçersiz fotoğraf alanı.")
        return redirect(hedef)

    foto=request.FILES.get("foto")
    ok,hata=_saha_foto_dogrula(foto)
    if not ok:
        messages.error(request,hata)
        return redirect(hedef)

    kayit,_=IsEmriFotograf.objects.update_or_create(
        is_emri=emri,sira=sira,
        defaults={"etiket":etiketler[sira-1],"foto":foto,"yukleyen":request.user},
    )
    IslemLogu.objects.create(
        talep=emri.talep,kullanici=request.user,islem="WORKORDER_PHOTO",
        aciklama=f"{emri.is_emri_no}: {kayit.etiket} görseli kaydedildi.",
        varlik_turu="IsEmriFotograf",varlik_id=str(kayit.pk),
    )
    messages.success(request,f"{kayit.etiket} kaydedildi.")
    return redirect(hedef)


@login_required
def is_emri_detay(request,pk):
    emri=get_object_or_404(
        IsEmri.objects.select_related(
            "talep__ilce","talep__mahalle","talep__yol","talep__is_turu","talep__is_alt_turu",
            "talep__abone","talep__sorumlu_koordinator__kullanici",
            "atanan_saha__kullanici","atanan_saha__rol","olusturan"
        ).prefetch_related("talep__geri_bildirimler","talep__loglar","fotograflar"),
        pk=pk
    )
    if not request.user.is_superuser:
        p=get_profile(request.user)
        if not p:
            raise PermissionDenied
        if p.rol.panel_tipi=="saha" and emri.atanan_saha_id!=p.id:
            raise PermissionDenied
        if p.rol.panel_tipi=="sef" and not talep_erisim_var_mi(request.user,emri.talep):
            raise PermissionDenied

    t=emri.talep
    zaman=[]

    def ekle(tarih,baslik,aciklama="",tur="normal"):
        if tarih:
            zaman.append({
                "tarih":tarih,
                "baslik":baslik,
                "aciklama":aciklama,
                "tur":tur,
            })

    ekle(t.olusturulma_tarihi,"Talep Oluşturuldu",f"{t.talep_no} • {t.get_oncelik_display()}")
    hareketler=sorted(list(t.geri_bildirimler.all()),key=lambda x:x.tarih)
    koord=next((x for x in hareketler if x.durum=="sefe_gonderildi"),None)
    if koord:
        ekle(koord.tarih,"Şef / Koordinatör Kuyruğuna Düştü",koord.mesaj)

    ekle(emri.atama_tarihi,"İş Emri Sahaya Atandı",f"{emri.is_emri_no} • {emri.gonderen_birim}","assignment")
    ekle(emri.kabul_tarihi,"Saha İş Emrini Kabul Etti","Mobil/Web saha operasyonu başladı.")
    ekle(emri.yola_cikis_tarihi,"Saha Yola Çıktı","Ekip arıza noktasına hareket etti.")
    gps_text=""
    if emri.gps_dogrulama_tarihi:
        gps_text=(
            f"GPS: {'Doğrulandı' if emri.gps_dogrulandi else 'Kaydedildi'}"
            + (f" • İş emri konumuna {emri.gps_mesafe_m} m" if emri.gps_mesafe_m is not None else "")
        )
    ekle(emri.adrese_ulasma_tarihi,"Adrese Ulaşıldı",gps_text,"gps")
    ekle(
        emri.mudahale_baslama_tarihi,
        "Müdahaleye Başlandı",
        "İş emri görselleri operasyon süresince iş emrine eklenir.",
        "photo",
    )
    ekle(
        emri.saha_tamam_tarihi,
        "Saha İşi Bitirdi",
        emri.sonuc_notu or t.saha_sonuc_notu,
        "photo",
    )

    for log in sorted(
        [x for x in t.loglar.all() if x.islem=="CHIEF_RETURN"],
        key=lambda x:x.tarih
    ):
        ekle(log.tarih,"Şef Sahaya Geri Gönderdi",log.aciklama,"return")

    ekle(emri.sef_onay_tarihi,"Şef İş Emrini Onayladı","Teknik operasyon tamamlandı.","success")
    if t.vatandas_bildirim_tarihi:
        ekle(t.vatandas_bildirim_tarihi,"185 Vatandaş Geri Dönüşü","Vatandaş bilgilendirme süreci tamamlandı.","callback")
    zaman.sort(key=lambda x:x["tarih"])

    foto_slotlari,foto_yuklenen,foto_zorunlu=_is_emri_foto_durumu(emri)
    return render(request,"dashboard/is_emri_detay.html",{
        "emri":emri,
        "talep":t,
        "zaman_cizelgesi":zaman,
        "foto_slotlari":foto_slotlari,
        "foto_yuklenen":foto_yuklenen,
        "foto_zorunlu":foto_zorunlu,
    })


@login_required
@panel_required("saha")
def mobil_saha(request):
    p=operational_profile(request,"saha")
    if _web_only_abone_role(p.rol):
        return redirect("dashboard:abone_islemleri")
    if not p.rol.mobil_erisim_var_mi():
        return redirect("dashboard:saha")

    # V47 — Şef onayına gönderilen iş artık saha ekibinin aktif/işlem yapılabilir
    # kuyruğunda tutulmaz. Şef geri gönderirse durum yeniden "sahaya_atandi" olur
    # ve iş otomatik olarak saha kuyruğuna geri gelir.
    qs=IsEmri.objects.filter(
        atanan_saha=p,
        talep__durum__in=["sahaya_atandi","kabul_edildi","yolda","yerinde","islemde"],
    ).select_related(
        "talep__ilce","talep__mahalle","talep__yol",
        "talep__is_turu","talep__is_alt_turu",
        "talep__sorumlu_koordinator__kullanici",
    ).prefetch_related("fotograflar")

    # V49 — saha iş listesi artık tek-aktif kuyruk mantığı kullanmaz.
    # Devam eden işler üstte, henüz kabul edilmemiş atanmış işler hemen altında
    # gösterilir; her atanmış iş bağımsız olarak kabul edilip ilerletilebilir.
    aktif_durumlar={"kabul_edildi","yolda","yerinde","islemde"}
    oncelik_sirasi={"acil":0,"yuksek":1,"normal":2,"dusuk":3}
    tum_isler=list(qs)
    tum_isler.sort(key=lambda e:(
        0 if e.talep.durum in aktif_durumlar else 1,
        oncelik_sirasi.get(e.talep.oncelik,9),
        e.atama_tarihi or timezone.now(),
    ))

    for e in tum_isler:
        slotlar,yuklenen,zorunlu=_is_emri_foto_durumu(e)
        e.foto_slotlari=slotlar
        e.foto_yuklenen=yuklenen
        e.foto_zorunlu=zorunlu
        if e.talep.durum in aktif_durumlar:
            e.kuyruk_etiketi="AKTİF İŞ"
            e.kuyruk_sinifi="active"
        elif e.talep.durum=="sahaya_atandi":
            e.kuyruk_etiketi="ATANMIŞ İŞ"
            e.kuyruk_sinifi="assigned"
        elif e.talep.durum=="onay_bekliyor":
            e.kuyruk_etiketi="ŞEF ONAYINDA"
            e.kuyruk_sinifi="approval"
        else:
            e.kuyruk_etiketi=e.get_durum_display().upper()
            e.kuyruk_sinifi="neutral"
    aktif_isler=tum_isler

    # V34 bildirim mantığı: sayaç "okundu" durumuna değil aktif iş emrine bağlıdır.
    # İş tamamlanıncaya kadar üstte görünür; Tamam yalnız modalı kapatır.
    modal_bildirimler=[]
    for emri in aktif_isler:
        b=MobilBildirim.objects.filter(kullanici=request.user,is_emri=emri).order_by("-olusturulma_tarihi").first()
        if b:
            baslik=b.baslik
            mesaj=b.mesaj
            tarih=b.olusturulma_tarihi
        else:
            t=emri.talep
            baslik=f"Aktif İş Emri {emri.is_emri_no}"
            mesaj=f"{t.ilce.ad} / {t.mahalle.ad} • {t.is_turu.ad} → {t.is_alt_turu.ad}"
            tarih=emri.atama_tarihi
        modal_bildirimler.append({
            "id":b.pk if b else None,
            "is_emri_id":emri.pk,
            "is_emri_no":emri.is_emri_no,
            "talep_no":emri.talep.talep_no,
            "baslik":baslik,
            "mesaj":mesaj,
            "tarih":tarih,
        })

    return render(request,"mobile/saha.html",{
        "profil":p,
        "is_emirleri":aktif_isler,
        "aktif":len(aktif_isler),
        "acil":sum(1 for e in aktif_isler if e.talep.oncelik=="acil"),
        "modal_bildirimler":modal_bildirimler,
        "aktif_bildirim_sayisi":len(aktif_isler),
    })


@require_GET
def mobil_bildirimler_api(request):
    user=request.user if request.user.is_authenticated else _mobil_token_user(request)
    if not user:
        return JsonResponse({"ok":False,"message":"Yetkisiz."},status=401)
    p=get_profile(user)
    if not p or p.rol.panel_tipi!="saha" or _web_only_abone_role(p.rol):
        return JsonResponse({"ok":False,"message":"Bu rol mobil saha bildirimlerine yetkili değil."},status=403)

    aktif_qs=IsEmri.objects.filter(
        atanan_saha=p,
        talep__durum__in=["sahaya_atandi","kabul_edildi","yolda","yerinde","islemde"],
    ).select_related(
        "talep__ilce","talep__mahalle","talep__is_turu","talep__is_alt_turu"
    ).order_by("-atama_tarihi")

    data=[]
    for emri in aktif_qs:
        b=MobilBildirim.objects.filter(kullanici=user,is_emri=emri).order_by("-olusturulma_tarihi").first()
        t=emri.talep
        data.append({
            "id":b.pk if b else f"work-{emri.pk}",
            "is_emri_id":emri.pk,
            "is_emri_no":emri.is_emri_no,
            "talep_no":t.talep_no,
            "baslik":b.baslik if b else f"Aktif İş Emri {emri.is_emri_no}",
            "mesaj":b.mesaj if b else f"{t.ilce.ad} / {t.mahalle.ad} • {t.is_turu.ad} → {t.is_alt_turu.ad}",
            "tarih":timezone.localtime(b.olusturulma_tarihi if b else emri.atama_tarihi).isoformat() if (b or emri.atama_tarihi) else None,
        })

    # unread geriye dönük Android uyumluluğu için tutulur; web arayüzü active_count kullanır.
    return JsonResponse({
        "ok":True,
        "active_count":len(data),
        "unread":MobilBildirim.objects.filter(kullanici=user,okundu=False).count(),
        "bildirimler":data,
    })


@csrf_exempt
@require_POST
def mobil_bildirim_okundu(request,pk):
    user=request.user if request.user.is_authenticated else _mobil_token_user(request)
    if not user:
        return JsonResponse({"ok":False,"message":"Yetkisiz."},status=401)
    b=get_object_or_404(MobilBildirim,pk=pk,kullanici=user)
    if not b.okundu:
        b.okundu=True
        b.save(update_fields=["okundu"])
    return JsonResponse({"ok":True})


def _mobil_token_user(request):
    auth=(request.headers.get("Authorization") or "").strip()
    if not auth.lower().startswith("bearer "):
        return None
    key=auth.split(" ",1)[1].strip()
    token=MobilToken.objects.select_related("kullanici__personel_profili__rol").filter(
        anahtar=key,aktif=True,kullanici__is_active=True
    ).first()
    if not token:
        return None
    token.son_kullanim=timezone.now()
    token.save(update_fields=["son_kullanim"])
    return token.kullanici


@csrf_exempt
@require_POST
def mobil_api_giris(request):
    try:
        payload=json.loads(request.body.decode("utf-8") or "{}")
    except (ValueError,UnicodeDecodeError):
        return JsonResponse({"ok":False,"message":"Geçersiz JSON."},status=400)
    username=(payload.get("username") or "").strip()
    password=payload.get("password") or ""
    user=authenticate(request,username=username,password=password)
    if not user or not user.is_active:
        return JsonResponse({"ok":False,"message":"Kullanıcı adı veya şifre hatalı."},status=401)
    p=get_profile(user)
    if (
        not p or not p.aktif or p.rol.panel_tipi!="saha"
        or _web_only_abone_role(p.rol)
        or not p.rol.mobil_erisim_var_mi()
    ):
        return JsonResponse(
            {"ok":False,"message":"Bu rol Mobil Saha çalışma kanalına yetkili değil."},
            status=403,
        )

    token=MobilToken.objects.filter(kullanici=user).first()
    if not token:
        token=MobilToken.objects.create(kullanici=user,anahtar=secrets.token_hex(24))
    elif not token.aktif:
        token.aktif=True
        token.anahtar=secrets.token_hex(24)
        token.save(update_fields=["aktif","anahtar"])
    return JsonResponse({
        "ok":True,
        "token":token.anahtar,
        "user":{
            "username":user.username,
            "name":user.get_full_name() or user.username,
            "role":p.rol.ad,
        },
    })


@require_GET
def mobil_api_is_emirleri(request):
    user=_mobil_token_user(request)
    if not user:
        return JsonResponse({"ok":False,"message":"Yetkisiz."},status=401)
    p=get_profile(user)
    if (
        not p or p.rol.panel_tipi!="saha"
        or _web_only_abone_role(p.rol)
        or not p.rol.mobil_erisim_var_mi()
    ):
        return JsonResponse({"ok":False,"message":"Bu rol Mobil Saha kanalına yetkili değil."},status=403)
    qs=IsEmri.objects.filter(
        atanan_saha=p,
        talep__durum__in=["sahaya_atandi","kabul_edildi","yolda","yerinde","islemde"],
    ).select_related(
        "talep__ilce","talep__mahalle","talep__yol",
        "talep__is_turu","talep__is_alt_turu",
    ).prefetch_related("fotograflar").order_by("-atama_tarihi")
    data=[]
    for e in qs:
        t=e.talep
        data.append({
            "id":e.pk,
            "is_emri_no":e.is_emri_no,
            "talep_no":t.talep_no,
            "durum":e.durum,
            "durum_label":e.get_durum_display(),
            "oncelik":t.oncelik,
            "oncelik_label":t.get_oncelik_display(),
            "is_turu":t.is_turu.ad,
            "is_alt_turu":t.is_alt_turu.ad,
            "aciklama":t.aciklama,
            "adres":{
                "ilce":t.ilce.ad,
                "mahalle":t.mahalle.ad,
                "yol":t.yol.ad,
                "kapi_no":t.kapi_no,
                "tarif":t.adres_aciklama,
                "lat":float(t.lat) if t.lat is not None else None,
                "lng":float(t.lng) if t.lng is not None else None,
            },
            "atama_tarihi":timezone.localtime(e.atama_tarihi).isoformat() if e.atama_tarihi else None,
            "zorunlu_fotograflar":[
                {"sira":x["sira"],"etiket":x["etiket"],"tamam":x["tamam"]}
                for x in _is_emri_foto_slotlari(e)
            ],
        })
    return JsonResponse({"ok":True,"count":len(data),"is_emirleri":data})


@csrf_exempt
@require_POST
def mobil_api_durum(request,pk):
    user=_mobil_token_user(request)
    if not user:
        return JsonResponse({"ok":False,"message":"Yetkisiz."},status=401)
    p=get_profile(user)
    if (
        not p or p.rol.panel_tipi!="saha"
        or _web_only_abone_role(p.rol)
        or not p.rol.mobil_erisim_var_mi()
    ):
        return JsonResponse({"ok":False,"message":"Bu rol Mobil Saha kanalına yetkili değil."},status=403)
    emri=get_object_or_404(IsEmri.objects.select_related("talep__is_alt_turu").prefetch_related("fotograflar"),pk=pk,atanan_saha=p)
    try:
        payload=json.loads(request.body.decode("utf-8") or "{}")
    except (ValueError,UnicodeDecodeError):
        return JsonResponse({"ok":False,"message":"Geçersiz JSON."},status=400)

    yeni=(payload.get("durum") or "").strip()
    t=emri.talep
    gecisler={
        "sahaya_atandi":"kabul_edildi",
        "kabul_edildi":"yolda",
        "yolda":"yerinde",
        "yerinde":"islemde",
        "islemde":"onay_bekliyor",
    }
    if gecisler.get(t.durum)!=yeni:
        return JsonResponse({"ok":False,"message":"İş akışı sırası geçersiz."},status=409)

    # V49 — Mobil API de web arayüzüyle aynı şekilde çoklu aktif işi destekler.
    # Başka bir iş devam ediyor olsa bile atanmış yeni iş kabul edilebilir.

    if yeni=="onay_bekliyor":
        slotlar,yuklenen,zorunlu=_is_emri_foto_durumu(emri)
        eksikler=[x["etiket"] for x in slotlar if not x["tamam"]]
        if eksikler:
            return JsonResponse({
                "ok":False,
                "message":f"Bu iş alt türü için {zorunlu} zorunlu fotoğraf gerekiyor. Eksik: {', '.join(eksikler)}.",
                "zorunlu_fotograf":zorunlu,
                "yuklenen_fotograf":yuklenen,
                "eksikler":eksikler,
            },status=400)
        not_metni=(payload.get("sonuc_notu") or "").strip()
        if not not_metni:
            return JsonResponse({"ok":False,"message":"Şef onayı için sonuç notu zorunludur."},status=400)
        t.saha_sonuc_notu=not_metni[:1000]
        t.saha_tamam_bildirim_tarihi=timezone.now()
        t.sef_onaylayan=None
        t.sef_onay_tarihi=None

    eski=t.durum
    t.durum=yeni
    t.save()
    emri=_is_emri_esitle(t,user)
    GeriBildirim.objects.create(
        talep=t,kullanici=user,durum=yeni,sistem_mesaji=True,
        mesaj=f"Mobil saha uygulamasından iş emri durumu güncellendi: {emri.get_durum_display()}."
    )
    IslemLogu.objects.create(
        talep=t,kullanici=user,islem="MOBILE_FIELD_STATUS",
        aciklama=f"{emri.is_emri_no} mobil uygulamadan {eski} → {yeni} güncellendi.",
        varlik_turu="IsEmri",varlik_id=str(emri.pk),
        eski_deger=eski,yeni_deger=yeni,
    )
    return JsonResponse({
        "ok":True,
        "is_emri_no":emri.is_emri_no,
        "durum":emri.durum,
        "durum_label":emri.get_durum_display(),
    })


@csrf_exempt
@require_POST
def mobil_api_fotograf(request,pk):
    user=_mobil_token_user(request)
    if not user:
        return JsonResponse({"ok":False,"message":"Yetkisiz."},status=401)
    p=get_profile(user)
    if not p or p.rol.panel_tipi!="saha" or _web_only_abone_role(p.rol) or not p.rol.mobil_erisim_var_mi():
        return JsonResponse({"ok":False,"message":"Bu rol Mobil Saha kanalına yetkili değil."},status=403)
    emri=get_object_or_404(
        IsEmri.objects.select_related("talep__is_alt_turu"),pk=pk,atanan_saha=p
    )
    if emri.durum not in {"yerinde","islemde"}:
        return JsonResponse({"ok":False,"message":"Görseller adrese ulaşıldıktan sonra eklenebilir."},status=409)
    try:
        sira=int(request.POST.get("sira") or 0)
    except ValueError:
        sira=0
    etiketler=emri.talep.is_alt_turu.zorunlu_foto_etiketleri
    if not 1<=sira<=len(etiketler):
        return JsonResponse({"ok":False,"message":"Geçersiz fotoğraf alanı."},status=400)
    foto=request.FILES.get("foto")
    ok,hata=_saha_foto_dogrula(foto)
    if not ok:
        return JsonResponse({"ok":False,"message":hata},status=400)
    kayit,_=IsEmriFotograf.objects.update_or_create(
        is_emri=emri,sira=sira,
        defaults={"etiket":etiketler[sira-1],"foto":foto,"yukleyen":user},
    )
    return JsonResponse({
        "ok":True,"sira":kayit.sira,"etiket":kayit.etiket,
        "zorunlu_fotograf":len(etiketler),"yuklenen_fotograf":emri.fotograflar.count(),
    })


@require_GET
def mobil_service_worker(request):
    response=HttpResponse(
        'const CACHE="isu185-saha-v28";self.addEventListener("install",()=>self.skipWaiting());self.addEventListener("activate",e=>e.waitUntil(self.clients.claim()));self.addEventListener("fetch",()=>{});',
        content_type="application/javascript; charset=utf-8",
    )
    response["Service-Worker-Allowed"]="/"
    response["Cache-Control"]="no-cache"
    return response
